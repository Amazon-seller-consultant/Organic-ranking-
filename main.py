import os
import csv
import io
import asyncio
import time
import logging
import math
from datetime import datetime, timezone
from uuid import uuid4
import httpx
import anthropic
from pathlib import Path
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, JSONResponse
from pydantic import BaseModel
from typing import Optional

load_dotenv(Path(__file__).parent / ".env", override=True)

app = FastAPI(title="Amazon Rank Tracker")
logger = logging.getLogger(__name__)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

SERPAPI_KEY = os.getenv("SERPAPI_KEY", "")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
CANOPY_API_KEY = os.getenv("CANOPY_API_KEY", "")

SERPAPI_BASE = "https://serpapi.com/search"
MAX_PAGES = 3
PAGE_SIZE = 16  # Amazon typically shows 16-24 results per page
SERPAPI_CONCURRENCY = max(1, int(os.getenv("SERPAPI_CONCURRENCY", "6")))
RANK_CACHE_TTL_SECONDS = max(0, int(os.getenv("RANK_CACHE_TTL_SECONDS", "21600")))
AI_CONCURRENCY = max(1, int(os.getenv("AI_CONCURRENCY", "2")))
CANOPY_CONCURRENCY = max(1, int(os.getenv("CANOPY_CONCURRENCY", "5")))
CANOPY_MAX_PAGES = max(1, int(os.getenv("CANOPY_MAX_PAGES", "3")))
CANOPY_SEARCH_URL = "https://rest.canopyapi.co/api/amazon/search"

_rank_cache: dict[tuple[str, str, str], tuple[float, dict]] = {}


@app.exception_handler(Exception)
async def unhandled_exception_handler(request, exc):
    logger.exception("Unhandled error for %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=500,
        content={
            "detail": (
                f"Internal server error ({type(exc).__name__}). "
                "Please retry once; if it continues, check the server logs."
            )
        },
    )


def _anthropic_create_message(client: anthropic.Anthropic, **kwargs):
    """Retry temporary Anthropic failures and return a useful API error."""
    last_error = None
    for attempt in range(3):
        try:
            return client.messages.create(**kwargs)
        except Exception as exc:
            last_error = exc
            status = getattr(exc, "status_code", None)
            retryable = status in {429, 500, 502, 503, 529} or status is None
            if retryable and attempt < 2:
                time.sleep(1.5 * (2 ** attempt))
                continue
            message = str(exc).strip() or type(exc).__name__
            raise HTTPException(
                status_code=502,
                detail=f"Anthropic API error: {message[:500]}",
            ) from exc
    raise HTTPException(
        status_code=502,
        detail=f"Anthropic API error: {str(last_error)[:500]}",
    )


class AsinRequest(BaseModel):
    asin: str


class SuggestRequest(BaseModel):
    asin: str
    product_title: str


class RankingRequest(BaseModel):
    asin: str
    terms: list[str]
    marketplace: Optional[str] = "amazon.com"


class PPCKeyword(BaseModel):
    keyword: str
    match_type: str
    campaign: str
    impressions: float
    top_of_search_is: Optional[float] = None
    clicks: float
    spend: float
    ctr: float
    orders: float
    cvr: float
    sales: float
    acos: Optional[float] = None
    cpc: float
    bid: float
    suggested_bid: Optional[float] = None
    roas: Optional[float] = None
    organic_rank: Optional[int] = None


class PPCAnalyzeRequest(BaseModel):
    keywords: list[PPCKeyword]
    target_acos: Optional[float] = None


class PVOMatchRow(BaseModel):
    match_type: str
    campaign: str
    ad_group: Optional[str] = None
    impressions: float
    clicks: float
    ctr: float
    cpc: float
    spend: float
    orders: float
    sales: float
    cvr: float
    acos: Optional[float] = None
    roas: Optional[float] = None
    top_of_search_is: Optional[float] = None


class PVOKeyword(BaseModel):
    keyword: str
    organic_rank: int
    ppc_rows: list[PVOMatchRow]


class PVORequest(BaseModel):
    asin: str
    keywords: list[PVOKeyword]
    target_acos: Optional[float] = None


def check_api_keys():
    if not SERPAPI_KEY:
        raise HTTPException(status_code=500, detail="SERPAPI_KEY not configured. Add it to your .env file.")
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured. Add it to your .env file.")


@app.get("/")
async def root():
    return FileResponse("static/index.html")


@app.post("/api/product-info")
async def get_product_info(req: AsinRequest):
    if not SERPAPI_KEY:
        raise HTTPException(status_code=500, detail="SERPAPI_KEY not configured.")

    asin = req.asin.strip().upper()
    if not asin:
        raise HTTPException(status_code=400, detail="ASIN cannot be empty.")

    params = {
        "engine": "amazon_product",
        "asin": asin,
        "amazon_domain": "amazon.com",
        "api_key": SERPAPI_KEY,
    }

    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(SERPAPI_BASE, params=params)

    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"SerpAPI error: {resp.text[:200]}")

    data = resp.json()

    if "error" in data:
        raise HTTPException(status_code=400, detail=data["error"])

    product = data.get("product_results", {})
    if not product:
        raise HTTPException(status_code=404, detail=f"No product found for ASIN: {asin}")

    return {
        "asin": asin,
        "title": product.get("title", "Unknown Product"),
        "brand": product.get("brand", ""),
        "image": product.get("media", [{}])[0].get("link", "") if product.get("media") else product.get("image", ""),
        "rating": product.get("rating", None),
        "ratings_total": product.get("ratings_total", None),
        "price": product.get("price", {}).get("current", {}).get("value", None) if isinstance(product.get("price"), dict) else None,
    }


@app.post("/api/suggest-terms")
async def suggest_terms(req: SuggestRequest):
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured.")

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    prompt = f"""You are an Amazon SEO expert. Given this Amazon product, suggest 10 highly relevant search terms that real customers would type into Amazon's search bar to find this product.

Product ASIN: {req.asin}
Product Title: {req.product_title}

Rules:
- Return ONLY the search terms, one per line, no numbering, no explanation
- Return exactly this mix:
  - 2 short head terms (1-2 words)
  - 5 mid-tail terms (2-4 words)
  - 3 long-tail terms (4-6 words)
- Prioritize mid-tail terms because they are usually the best balance of search volume, buyer intent, and ranking opportunity
- Mid-tail terms should sound like normal Amazon searches, not overly specific phrases
- Think about what a real buyer would search for
- Terms should be 1-6 words each
- Avoid making every term a long-tail use-case phrase
- Do not include the ASIN itself"""

    message = _anthropic_create_message(
        client,
        model="claude-sonnet-4-6",
        max_tokens=500,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = message.content[0].text.strip()
    terms = [line.strip() for line in raw.splitlines() if line.strip()]
    # Remove any accidental numbering (1. 2. etc)
    cleaned = []
    for t in terms:
        if t and len(t) > 1:
            # Strip leading "1." or "1)" patterns
            if len(t) > 2 and t[0].isdigit() and t[1] in ".):":
                t = t[2:].strip()
            elif len(t) > 3 and t[0].isdigit() and t[1].isdigit() and t[2] in ".):":
                t = t[3:].strip()
            if t:
                cleaned.append(t)

    return {"terms": cleaned[:10]}


async def search_term_ranking(client: httpx.AsyncClient, asin: str, term: str, marketplace: str) -> dict:
    """Search Amazon for a keyword and find the ASIN's position."""
    asin = asin.upper()
    term = term.strip()
    cache_key = (marketplace.lower(), asin, term.lower())
    cached = _rank_cache.get(cache_key)
    if cached and time.monotonic() - cached[0] < RANK_CACHE_TTL_SECONDS:
        return dict(cached[1])

    position_overall = None
    found_page = None

    for page_num in range(1, MAX_PAGES + 1):
        params = {
            "engine": "amazon",
            "k": term,
            "amazon_domain": marketplace,
            "page": page_num,
            "api_key": SERPAPI_KEY,
        }

        try:
            resp = await client.get(SERPAPI_BASE, params=params, timeout=30)
            if resp.status_code != 200:
                raise HTTPException(
                    status_code=502,
                    detail=f"SerpAPI error: {resp.text[:300] or f'HTTP {resp.status_code}'}",
                )

            try:
                data = resp.json()
            except Exception:
                raise HTTPException(
                    status_code=502,
                    detail=f"SerpAPI returned an invalid response: {resp.text[:300]}",
                )
            if "error" in data:
                raise HTTPException(status_code=502, detail=f"SerpAPI error: {data['error']}")

            organic = data.get("organic_results", [])
            if not organic:
                break

            for index, item in enumerate(organic, 1):
                item_asin = (item.get("asin") or "").upper()
                if item_asin == asin:
                    pos_in_page = item.get("position", index)
                    position_overall = (page_num - 1) * PAGE_SIZE + pos_in_page
                    found_page = page_num
                    break

            if position_overall is not None:
                break

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"SerpAPI request failed: {e}")

    result = {
        "term": term,
        "position": position_overall,
        "page": found_page,
        "found": position_overall is not None,
    }
    if RANK_CACHE_TTL_SECONDS:
        _rank_cache[cache_key] = (time.monotonic(), result)
    return dict(result)


async def _rank_entries(entries: list[dict]) -> list[dict]:
    jobs = [
        (entry_index, term_index, entry["asin"], entry["marketplace"], term)
        for entry_index, entry in enumerate(entries)
        for term_index, term in enumerate(entry["terms"])
    ]
    grouped = [[None] * len(entry["terms"]) for entry in entries]
    unique_jobs = {}
    for _, _, asin, marketplace, term in jobs:
        key = (marketplace.lower(), asin.upper(), term.strip().lower())
        unique_jobs.setdefault(key, (asin, marketplace, term))

    limits = httpx.Limits(
        max_connections=max(SERPAPI_CONCURRENCY * 2, 10),
        max_keepalive_connections=max(SERPAPI_CONCURRENCY, 5),
    )
    semaphore = asyncio.Semaphore(SERPAPI_CONCURRENCY)

    async def run_limited(asin: str, marketplace: str, term: str) -> dict:
        async with semaphore:
            return await search_term_ranking(client, asin, term, marketplace)

    async with httpx.AsyncClient(limits=limits, timeout=35) as client:
        unique_rankings = await asyncio.gather(*[
            run_limited(asin, marketplace, term)
            for asin, marketplace, term in unique_jobs.values()
        ])
    ranking_by_key = dict(zip(unique_jobs, unique_rankings))
    for entry_index, term_index, asin, marketplace, term in jobs:
        key = (marketplace.lower(), asin.upper(), term.strip().lower())
        grouped[entry_index][term_index] = dict(ranking_by_key[key])
    return [
        {
            "asin": entry["asin"],
            "marketplace": entry["marketplace"],
            "rankings": grouped[index],
        }
        for index, entry in enumerate(entries)
    ]


@app.post("/api/check-rankings")
async def check_rankings(req: RankingRequest):
    if not CANOPY_API_KEY and not SERPAPI_KEY:
        raise HTTPException(status_code=500, detail="Neither CANOPY_API_KEY nor SERPAPI_KEY is configured.")
    if not req.terms:
        raise HTTPException(status_code=400, detail="No search terms provided.")

    asin = req.asin.strip().upper()
    marketplace = req.marketplace or "amazon.com"
    entries = [{
        "asin": asin,
        "marketplace": marketplace,
        "terms": [term.strip() for term in req.terms if term.strip()],
    }]
    return {"asin": asin, "results": await _canopy_rank_entries(entries)}


PPC_SYSTEM_PROMPT = """You are a senior Amazon PPC manager with deep expertise in organic ranking strategy, listing optimization, and advertising efficiency. You think strategically — never giving generic advice.

ORGANIC POSITION INTERPRETATION:
- Ranks 1-16: Strong visibility. Defend aggressively. Maintain PPC support carefully.
- Ranks 17-30: Bottom/mid Page 1. Strong scaling opportunity. PPC can significantly improve ranking.
- Ranks 31-48: Early Page 2. High opportunity zone. Push IF conversion metrics justify it.
- Ranks 49-80: Weak visibility. Analyze listing quality before scaling.
- Ranks 80+: Poor visibility. Check relevancy/indexing before increasing spend.

DECISION RULES:
1. Organic rank 20-45 + strong CVR + decent CTR + low impressions → increase bids moderately, increase Top of Search multiplier. Visibility is the bottleneck. Page 2→1 ranking opportunity.
2. High impressions + low CTR → do NOT recommend aggressive bid increases. Diagnose: weak main image, weak title, pricing mismatch, low reviews, weak differentiation, irrelevant targeting. Say: "Traffic exists already, but shoppers are not clicking."
3. Good CTR + poor CVR → diagnose: listing mismatch, weak bullets, poor A+ content, pricing issue, review concerns, mismatch between keyword intent and product page.
4. High ACoS + strong CVR + organic rank improving or near Page 1 → temporary high ACoS may be strategically acceptable. Explain ranking momentum logic. Recommend monitoring profitability. Do NOT automatically suggest lowering bids.
5. Low impressions + low clicks + poor organic ranking → check indexing/relevancy first. Optimize listing before scaling spend. Increasing bids alone may not solve visibility.
6. Strong CTR + strong CVR + low spend → scale aggressively. Increase Top of Search placement. Mention keyword expansion opportunity.
7. Very high spend + no conversions → recommend bid reduction or pause. Add negatives if applicable. Mention wasted spend risk.

TARGET ACoS LOGIC: If target ACoS provided, compare actual vs target. Do NOT blindly recommend lowering bids when ACoS exceeds target. Evaluate organic ranking opportunity, conversion quality, traffic potential, and ranking momentum first.

STRATEGIC UNDERSTANDING:
- Amazon organic ranking is influenced by sales velocity
- PPC supports organic rank growth
- Top of Search placements matter heavily
- Page 1 ranking has disproportionate value
- Strong CVR often justifies scaling even at higher ACoS
- Low CTR usually indicates listing issues, not PPC issues
- Analyze visibility and conversion separately

NEVER:
- Give generic recommendations
- Repeat identical suggestions across keywords
- Suggest lowering bids solely because ACoS is high
- Suggest increasing bids without explaining why
- Ignore organic ranking context
- Ignore CTR/CVR relationship

Every recommendation must reference specific metrics from the data provided."""


def format_keyword_for_analysis(kw: PPCKeyword, target_acos: Optional[float]) -> str:
    acos_str = f"{kw.acos * 100:.1f}%" if kw.acos is not None else "N/A (no sales)"
    target_str = f" [Target: {target_acos:.1f}%]" if target_acos else ""
    tos_str = f"{kw.top_of_search_is:.1f}%" if kw.top_of_search_is else "N/A"
    organic_str = f"Rank #{kw.organic_rank}" if kw.organic_rank else "Not available"
    roas_str = f"{kw.roas:.2f}x" if kw.roas else "N/A"
    return f"""Keyword: {kw.keyword} ({kw.match_type})
Campaign: {kw.campaign}
Organic Rank: {organic_str}
Impressions: {int(kw.impressions):,} | Top-of-Search IS: {tos_str}
Clicks: {int(kw.clicks)} | CTR: {kw.ctr * 100:.2f}%
Spend: ${kw.spend:.2f} | CPC: ${kw.cpc:.2f} | Bid: ${kw.bid:.2f}{f" | Suggested: ${kw.suggested_bid:.2f}" if kw.suggested_bid else ""}
Orders: {int(kw.orders)} | CVR: {kw.cvr:.1f}% | Sales: ${kw.sales:.2f}
ACoS: {acos_str}{target_str} | ROAS: {roas_str}"""


BATCH_SIZE = 7  # ~6000 tokens output per batch, well within 8192 limit


def _make_batch_prompt(keywords: list, target_acos: Optional[float]) -> str:
    formatted = "\n\n---\n\n".join(
        f"[KEYWORD {i+1}]\n{format_keyword_for_analysis(kw, target_acos)}"
        for i, kw in enumerate(keywords)
    )
    return f"""Analyze the following {len(keywords)} Amazon PPC keywords. For each provide a complete strategic analysis.

Return a JSON array — one object per keyword — with EXACTLY these fields:
"keyword", "match_type", "organic_strength", "traffic_opportunity", "conversion_quality",
"ppc_efficiency", "scaling_potential", "listing_weakness", "wasted_spend",
"ranking_momentum", "diagnosis", "recommended_action", "priority"

All string fields: 1-2 concise sentences. "diagnosis": 2-3 sentences. "recommended_action": specific steps referencing actual metric values.
"priority": exactly "High", "Medium", or "Low".
"listing_weakness" / "wasted_spend": use "No signals detected" / "No wasted spend detected" if not applicable.

Return ONLY a valid JSON array. No markdown fences. No text outside the array.

KEYWORDS:

{formatted}"""


def _parse_raw_json(raw: str) -> list:
    import json
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.rsplit("```", 1)[0].strip()
    return json.loads(raw)


@app.post("/api/ppc-analyze")
async def ppc_analyze(req: PPCAnalyzeRequest):
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured.")
    if not req.keywords:
        raise HTTPException(status_code=400, detail="No keywords provided.")
    if len(req.keywords) > 30:
        raise HTTPException(status_code=400, detail="Maximum 30 keywords per analysis.")

    batches = [req.keywords[i:i + BATCH_SIZE] for i in range(0, len(req.keywords), BATCH_SIZE)]
    semaphore = asyncio.Semaphore(AI_CONCURRENCY)

    def run_batch(batch: list[PPCKeyword]) -> list:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        message = _anthropic_create_message(
            client,
            model="claude-sonnet-4-6",
            max_tokens=7000,
            system=PPC_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": _make_batch_prompt(batch, req.target_acos)}],
        )
        raw = message.content[0].text.strip()
        try:
            return _parse_raw_json(raw)
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to parse AI response for batch: {e}. Raw: {raw[:300]}"
            )

    async def run_limited(batch: list[PPCKeyword]) -> list:
        async with semaphore:
            return await asyncio.to_thread(run_batch, batch)

    batch_results = await asyncio.gather(*[run_limited(batch) for batch in batches])
    all_results = [result for results in batch_results for result in results]
    return {"results": all_results}


PVO_SYSTEM_PROMPT = """You are a senior Amazon growth strategist. You analyze the SAME ASIN + keyword across multiple campaigns, ad groups, and match types simultaneously.

Your goal: determine which campaign structure performs best, which drives organic rank growth, and which wastes spend. You think like a ranking velocity expert, not just an ACoS manager.

ORGANIC RANK LOGIC:
- Ranks 1-16: Strong visibility. Defend aggressively. Maintain ranking dominance.
- Ranks 17-30: Strong Page 1 scaling opportunity. PPC can significantly improve rank.
- Ranks 31-48: Early Page 2. High ranking acceleration zone. Good candidates for aggressive scaling IF conversion metrics support it.
- Ranks 49-80: Weak visibility. Check listing quality before aggressive scaling.
- Ranks 80+: Poor ranking. Check relevancy/indexing before increasing spend.

CAMPAIGN & AD GROUP COMPARISON RULES:
- NEVER analyze campaigns independently. Always compare ALL campaigns/ad groups targeting the same keyword.
- Identify: highest CTR campaign, highest CVR campaign, best ACoS campaign, strongest sales velocity, wasted spend campaigns.
- The best campaign is NOT always lowest ACoS or cheapest CPC. It may be highest sales velocity, strongest CVR, or strongest organic ranking support.
- When one campaign dominates sales + CVR + ranking support → consolidate budget there, reduce inefficient overlap.
- Multiple campaigns targeting same keyword but only one converts → consolidate budget to winner, avoid internal competition.
- Highlight which specific campaign + ad group combination is the winning structure.

BOTTLENECK DIAGNOSIS RULES:
- High impressions + low CTR → listing issue (weak main image, weak title, pricing, low reviews). State: "Traffic exists but shoppers are not clicking."
- Good CTR + poor CVR → listing mismatch (bullets, A+ content, pricing, review concerns, keyword intent mismatch).
- Low impressions + low clicks + poor organic ranking → indexing/relevancy issue. Fix listing before scaling spend.
- Strong CTR + strong CVR + high ACoS + near Page 1 → acceptable. PPC supporting ranking momentum. Do NOT recommend lowering bids.
- High spend + no conversions → pause or reduce. Add negatives. Mention wasted spend risk.
- Strong CTR + strong CVR + low spend → scale aggressively. Increase Top of Search placement.

MATCH TYPE ANALYSIS:
- Exact with best CTR + CVR + lowest ACoS = prioritize, redirect budget from weaker types.
- Broad with high spend + weak CVR = reduce bids, keep only for harvesting.
- If only one match type converts = pause or reduce non-converting types.
- When organic rank 31-48 AND Exact CVR strong AND impressions low = raise Exact bids + Top of Search modifier.

TARGET ACoS LOGIC: If target ACoS provided, compare actual vs target. Do NOT blindly recommend lowering bids when ACoS exceeds target. Weigh ranking opportunity, conversion quality, and sales velocity first.

RANKING ACCELERATION SCORE RULES:
- HIGH: Strong CVR + rank 17-48 + adequate impressions + at least one dominant campaign → very likely to improve organic rank quickly with scaling.
- MEDIUM: Moderate CVR + Page 2 + some campaign confusion or inefficiency → possible improvement with restructuring.
- LOW: Poor CVR OR rank 80+ OR listing issues OR indexing problems → unlikely without foundational fixes first.

CRITICAL RULES — NEVER:
- Analyze campaigns independently without comparison.
- Recommend lowering bids solely because ACoS is high.
- Recommend increasing bids without reasoning.
- Ignore organic ranking context.
- Ignore campaign overlap.
- Ignore match type differences.
- Give generic advice that doesn't reference specific metric values.

Every recommendation must connect: organic rank + PPC efficiency + campaign structure + targeting quality + conversion quality + ranking acceleration potential + sales velocity."""


def _page_label(rank: int) -> str:
    if rank <= 16:   return "Page 1 (strong)"
    if rank <= 30:   return "Page 1 (lower)"
    if rank <= 48:   return "Page 2 (early)"
    if rank <= 80:   return "Page 2-3 (weak)"
    return "Page 3+ (poor)"


def _fmt_pvo_keyword(kw: PVOKeyword, target_acos: Optional[float]) -> str:
    lines = [
        f"KEYWORD: {kw.keyword}",
        f"ORGANIC RANK: #{kw.organic_rank} — {_page_label(kw.organic_rank)}",
    ]
    if target_acos:
        lines.append(f"TARGET ACoS: {target_acos:.1f}%")
    lines.append(f"TOTAL CAMPAIGN ENTRIES FOR THIS KEYWORD: {len(kw.ppc_rows)}")
    lines.append("")
    for i, row in enumerate(kw.ppc_rows, 1):
        tos = f"{row.top_of_search_is:.1f}%" if row.top_of_search_is else "N/A"
        acos_s = f"{row.acos * 100:.1f}%" if row.acos is not None else "N/A (no sales)"
        mt = row.match_type.replace("Keyword - ", "").replace("Product target - ", "").upper()
        ad_grp_line = f"\n  Ad Group: {row.ad_group}" if row.ad_group else ""
        lines += [
            f"[CAMPAIGN ENTRY {i}]",
            f"  Campaign: {row.campaign}{ad_grp_line}",
            f"  Match Type: {mt}",
            f"  Impressions: {int(row.impressions):,} | Top-of-Search IS: {tos} | CTR: {row.ctr * 100:.2f}%",
            f"  Clicks: {int(row.clicks)} | CPC: ${row.cpc:.2f} | Spend: ${row.spend:.2f}",
            f"  Orders: {int(row.orders)} | CVR: {row.cvr:.1f}% | Sales: ${row.sales:.2f} | ACoS: {acos_s}",
            "",
        ]
    return "\n".join(lines)


def _make_pvo_batch_prompt(keywords: list[PVOKeyword], target_acos: Optional[float]) -> str:
    bodies = "\n\n---\n\n".join(
        f"[KEYWORD {i+1}]\n{_fmt_pvo_keyword(kw, target_acos)}"
        for i, kw in enumerate(keywords)
    )
    return f"""Analyze {len(keywords)} keyword(s). For each, compare ALL campaigns and ad groups targeting the SAME ASIN + keyword. Think like a senior Amazon growth strategist managing ranking velocity.

Return a JSON array with one object per keyword. Each object MUST have EXACTLY these fields:
- "keyword": string
- "organic_rank": number
- "page_placement": string (e.g. "Page 2 (early)")
- "organic_position_summary": string (2-3 sentences: ranking strength, ranking opportunity, competitive context)
- "campaign_ad_group_comparison": string (compare ALL campaign entries — which has best CTR, CVR, ACoS, sales velocity, ranking support — reference specific numbers from the data. Identify the winner and the wasted spend.)
- "winning_structure": string (name the specific campaign + ad group + match type that performs best, explain why it should receive more budget, and how to scale it)
- "bottleneck_analysis": string (diagnose the primary limitation: visibility/CTR/conversion/listing quality/targeting inefficiency/wasted spend — cite specific metrics)
- "recommended_actions": string (specific numbered steps with actual metric values — bid changes, budget shifts, structural changes, listing fixes if needed)
- "ranking_acceleration_score": "LOW" | "MEDIUM" | "HIGH"
- "priority": "High" | "Medium" | "Low"

"ranking_acceleration_score" rules:
- HIGH: Strong CVR + rank 17-48 + adequate impressions + dominant campaign structure → likely to improve organic rank quickly
- MEDIUM: Moderate CVR + Page 2 + some campaign inefficiency → possible with restructuring
- LOW: Poor CVR OR rank 80+ OR listing/indexing issues → unlikely without foundational fixes

Return ONLY valid JSON array. No markdown fences. No text outside the array.

DATA:

{bodies}"""


PVO_BATCH_SIZE = 3


@app.post("/api/ppc-vs-organic")
async def ppc_vs_organic(req: PVORequest):
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured.")
    if not req.keywords:
        raise HTTPException(status_code=400, detail="No keywords provided.")
    if len(req.keywords) > 30:
        raise HTTPException(status_code=400, detail="Maximum 30 keywords per analysis.")

    batches = [req.keywords[i:i + PVO_BATCH_SIZE] for i in range(0, len(req.keywords), PVO_BATCH_SIZE)]
    semaphore = asyncio.Semaphore(AI_CONCURRENCY)

    def run_batch(batch: list[PVOKeyword]) -> list:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        msg = _anthropic_create_message(
            client,
            model="claude-sonnet-4-6",
            max_tokens=7000,
            system=PVO_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": _make_pvo_batch_prompt(batch, req.target_acos)}],
        )
        raw = msg.content[0].text.strip()
        try:
            return _parse_raw_json(raw)
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to parse AI response: {e}. Raw: {raw[:300]}"
            )

    async def run_limited(batch: list[PVOKeyword]) -> list:
        async with semaphore:
            return await asyncio.to_thread(run_batch, batch)

    batch_results = await asyncio.gather(*[run_limited(batch) for batch in batches])
    all_results = [result for results in batch_results for result in results]
    return {"results": all_results, "asin": req.asin}


@app.get("/api/template.csv")
async def download_csv_template():
    rows = [
        ["ASIN", "Marketplace", "Search Term 1", "Search Term 2", "Search Term 3",
         "Search Term 4", "Search Term 5", "Search Term 6", "Search Term 7", "Search Term 8"],
        ["B08N5WRWNW", "amazon.com", "echo dot", "alexa speaker", "smart speaker",
         "echo dot 4th gen", "amazon echo", "voice assistant speaker", "", ""],
        ["B07XJ8C8F7", "amazon.com", "fire tv stick", "streaming device", "hdmi streaming",
         "4k fire stick", "amazon fire tv", "", "", ""],
    ]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=asin_rank_template.csv"},
    )


def parse_upload(content: bytes, filename: str) -> list[dict]:
    """Parse CSV or XLSX upload. Returns list of {asin, marketplace, terms}."""
    rows = []

    if filename.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        raw_rows = [[str(cell.value).strip() if cell.value is not None else "" for cell in row]
                    for row in ws.iter_rows()]
        wb.close()
    else:
        text = content.decode("utf-8-sig", errors="replace")
        raw_rows = [row for row in csv.reader(text.splitlines())]

    if not raw_rows:
        return rows

    # Skip header row
    data_rows = raw_rows[1:]

    merged: dict[tuple[str, str], list[str]] = {}
    seen_terms: dict[tuple[str, str], set[str]] = {}
    for row in data_rows:
        if not row:
            continue
        asin = row[0].strip().upper() if len(row) > 0 else ""
        if not asin or asin in ("ASIN", ""):
            continue
        marketplace = row[1].strip() if len(row) > 1 and row[1].strip() else "amazon.com"
        terms = [t.strip() for t in row[2:] if len(row) > 2 and t.strip()]
        if asin and terms:
            key = (asin, marketplace)
            merged.setdefault(key, [])
            seen_terms.setdefault(key, set())
            for term in terms:
                term_key = " ".join(term.lower().split())
                if term_key not in seen_terms[key]:
                    seen_terms[key].add(term_key)
                    merged[key].append(term)

    for (asin, marketplace), terms in merged.items():
        rows.append({"asin": asin, "marketplace": marketplace, "terms": terms})

    return rows


@app.post("/api/bulk-check")
async def bulk_check(file: UploadFile = File(...)):
    if not CANOPY_API_KEY and not SERPAPI_KEY:
        raise HTTPException(status_code=500, detail="Neither CANOPY_API_KEY nor SERPAPI_KEY is configured.")

    filename = file.filename or ""
    if not (filename.lower().endswith(".csv") or filename.lower().endswith((".xlsx", ".xlsm"))):
        raise HTTPException(status_code=400, detail="Only .csv, .xlsx, and .xlsm files are supported.")

    content = await file.read()
    try:
        entries = parse_upload(content, filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not entries:
        raise HTTPException(status_code=400, detail="No valid ASIN + search term rows found. Check the file format.")

    flat_results = await _canopy_rank_entries(entries)
    grouped = []
    offset = 0
    for entry in entries:
        count = len(entry["terms"])
        grouped.append({
            "asin": entry["asin"],
            "marketplace": entry["marketplace"],
            "rankings": flat_results[offset:offset + count],
        })
        offset += count
    return {"results": grouped}


# ─────────────────────────────────────────────────────────────────────────────
# Amazon Ads Opportunity Analyzer (full workbook + organic rank → campaign table)
# ─────────────────────────────────────────────────────────────────────────────

class ReportCampaignRow(BaseModel):
    campaign: str
    ad_group: str = ""
    match_type: str
    source: str = ""
    impressions: float
    ctr: float        # decimal (0.05 = 5%)
    cpc: float
    spend: float
    sales: float
    orders: float
    cvr: float        # decimal (0.05 = 5%)
    acos: Optional[float] = None   # decimal (0.2 = 20%)
    top_of_search_is: Optional[float] = None
    placement_top_acos: Optional[float] = None
    placement_top_orders: Optional[float] = None
    asin_count: Optional[int] = None
    confidence: Optional[str] = None


class ReportKeyword(BaseModel):
    asin: str
    keyword: str
    organic_rank: int
    campaigns: list[ReportCampaignRow]


class ReportAnalysisRequest(BaseModel):
    keywords: list[ReportKeyword]
    target_acos: float = 20.0   # percentage (20 = 20%)


class CSVExportRequest(BaseModel):
    filename: str
    rows: list[list[str]]


REPORT_SYSTEM_PROMPT = """You are a senior Amazon PPC strategist producing a concise decision table. For each ASIN + keyword/search query group, you compare ALL campaigns/ad groups and output ONE recommendation row per campaign entry.

You are analyzing a full Amazon Sponsored Products workbook with search term, targeting, placement, and advertised-product sheets. Some ASIN assignments are inferred from Campaign + Ad Group using the advertised-product sheet. Treat each campaign/ad group as a separate structure, compare them against each other, then choose winners.

ORGANIC RANK LOGIC:
- Ranks 1-16 (Page 1 strong): Defend. Scale winning campaigns +5% to +10%.
- Ranks 17-30 (Page 1 lower): Strong scaling opportunity. Good CVR → +10% to +20%.
- Ranks 31-48 (Page 2 early): Acceleration zone. Strong CVR → +15% to +30%.
- Ranks 49-80 (Page 2-3 weak): Check listing quality. Max +5% until listing improves.
- Ranks 80+ (Page 3+ poor): Relevancy/indexing issue. 0% or reduce.

BID CHANGE RULES:
- Rank 17-48 + CVR >5% + CTR decent + ACoS near or below target → +15% to +30%
- Rank 1-16 + strong CVR → +5% to +10% to defend
- Rank 17-48 + CVR weak (<3%) → +5% or 0% depending on CTR
- High impressions + low CTR (< 0.3%) → 0% (listing issue, not bid issue)
- ACoS > 2× target + weak CVR → -15% to -25%
- ACoS > 1.5× target + weak CVR → -10% to -15%
- Strong CVR + low impressions + rank 17-48 → +20% to +40% (visibility bottleneck)
- No orders + high spend → -20% or pause
- Broad match with poor CVR → "Reduce Broad Match Spend" with -10% to -20%

ACTION STRINGS (use EXACTLY one per row):
"Scale Aggressively" | "Moderate Increase" | "Hold" | "Reduce Waste" | "Improve Listing First" | "Focus Exact Match" | "Reduce Broad Match Spend"

CRITICAL RULES:
- Analyze ONLY ASIN + keyword/search query groups supplied by the user filter file
- When multiple campaigns target the same ASIN+keyword/search query, compare them explicitly before assigning actions
- The campaign with best CVR + best ACoS relative to target = winner → Scale Aggressively / Moderate Increase
- Other campaigns for same ASIN+keyword = Reduce Waste or Hold
- NEVER recommend bid increases when CTR < 0.3% and impressions > 5000 (listing issue)
- NEVER ignore organic rank context
- asin_count = number of ASINs sharing this campaign slot. When asin_count > 5 and confidence is LOW, treat data as less reliable."""


def _page_label_report(rank: int) -> str:
    if rank <= 16:  return "Page 1 (strong)"
    if rank <= 30:  return "Page 1 (lower)"
    if rank <= 48:  return "Page 2 (early)"
    if rank <= 80:  return "Page 2-3 (weak)"
    return "Page 3+ (poor)"


def _fmt_report_keyword(kw: ReportKeyword, target_acos: float) -> str:
    lines = [
        f"ASIN: {kw.asin}",
        f"KEYWORD: {kw.keyword}",
        f"ORGANIC RANK: #{kw.organic_rank} — {_page_label_report(kw.organic_rank)}",
        f"TARGET ACoS: {target_acos:.1f}%",
        f"TOTAL CAMPAIGN ENTRIES: {len(kw.campaigns)}",
        "",
    ]
    for i, c in enumerate(kw.campaigns, 1):
        acos_s = f"{c.acos * 100:.1f}%" if c.acos is not None else "N/A"
        cnt_s  = f" | asin_count={c.asin_count}" if c.asin_count else ""
        conf_s = f" | confidence={c.confidence}" if c.confidence else ""
        lines += [
            f"[CAMPAIGN {i}]",
            f"  Campaign: {c.campaign}",
            f"  Ad Group: {c.ad_group or 'N/A'}{cnt_s}{conf_s}",
            f"  Source: {c.source or 'N/A'} | Match Type: {c.match_type}",
            f"  Impressions: {int(c.impressions):,} | CTR: {c.ctr * 100:.3f}% | CPC: ${c.cpc:.2f}",
            f"  Orders: {int(c.orders)} | CVR: {c.cvr * 100:.2f}% | Spend: ${c.spend:.2f} | Sales: ${c.sales:.2f} | ACoS: {acos_s}",
            f"  Top-of-Search IS: {c.top_of_search_is * 100:.1f}%" if c.top_of_search_is is not None else "  Top-of-Search IS: N/A",
            f"  Top-of-Search Placement ACoS: {c.placement_top_acos * 100:.1f}% | Top-of-Search Orders: {int(c.placement_top_orders or 0)}" if c.placement_top_acos is not None else "  Top-of-Search Placement ACoS: N/A",
            "",
        ]
    return "\n".join(lines)


def _make_report_batch_prompt(keywords: list[ReportKeyword], target_acos: float) -> str:
    bodies = "\n\n---\n\n".join(
        f"[GROUP {i+1}]\n{_fmt_report_keyword(kw, target_acos)}"
        for i, kw in enumerate(keywords)
    )
    return f"""Analyze {len(keywords)} ASIN+keyword group(s). Compare ALL campaigns per group. Output ONE table row per campaign entry, plus one short ASIN-level summary per ASIN.

Return valid JSON with exactly this shape:
{{
  "rows": [
    {{
      "asin": "string",
      "keyword": "string",
      "organic_rank": 22,
      "amazon_page": "Page 1 (lower)",
      "campaign_name": "string",
      "ad_group_name": "string",
      "match_type": "EXACT",
      "source": "Search Term",
      "ctr": "1.23%",
      "cvr": "5.60%",
      "acos": "18.5%",
      "top_of_search": "12.5%",
      "target_acos": "20.0%",
      "bid_change": "+20%",
      "action": "Scale Aggressively",
      "reason": "One concise sentence specific to this campaign versus alternatives."
    }}
  ],
  "summaries": [
    {{
      "asin": "string",
      "best_campaign": "string",
      "wasted_campaign": "string",
      "best_keyword_opportunity": "string",
      "budget_shift": "string"
    }}
  ]
}}

Return ONLY valid JSON. No markdown fences.

DATA:

{bodies}"""


REPORT_BATCH_SIZE = 3


def _report_json_parts(raw: str) -> tuple[list, list]:
    parsed = _parse_raw_json(raw)
    if isinstance(parsed, list):
        return parsed, []
    if isinstance(parsed, dict):
        return parsed.get("rows", []) or [], parsed.get("summaries", []) or []
    return [], []


def _num(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace("%", "").replace("$", "").replace(",", "").strip())
    except Exception:
        return 0.0


def _rate(value) -> float:
    n = _num(value)
    return n / 100 if n > 1 else n


def _norm_text(value) -> str:
    return str(value or "").strip()


def _norm_key(value) -> str:
    return _norm_text(value).lower()


def _keyword_key(value) -> str:
    text = _norm_key(value).strip(" \t\r\n\"'“”‘’,,.;:")
    return " ".join(text.split())


def _find_sheet_name(wb, *needles: str) -> Optional[str]:
    normalized = {
        name: "".join(ch for ch in name.lower() if ch.isalnum())
        for name in wb.sheetnames
    }
    for name, compact in normalized.items():
        if all("".join(ch for ch in needle.lower() if ch.isalnum()) in compact for needle in needles):
            return name
    return None


def _worksheet_dicts(ws) -> list[dict]:
    if hasattr(ws, "reset_dimensions"):
        try:
            ws.reset_dimensions()
        except Exception:
            pass
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_text(h) for h in rows[0]]
    output = []
    for raw in rows[1:]:
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        output.append(row)
    return output


def _sheet_dicts(wb, sheet_name: str = "", *needles: str) -> list[dict]:
    actual_name = sheet_name if sheet_name in wb.sheetnames else _find_sheet_name(wb, *(needles or (sheet_name,)))
    if not actual_name:
        return []
    return _worksheet_dicts(wb[actual_name])


def _row_get(row: dict, *names: str):
    lower = {k.lower().strip(): v for k, v in row.items()}
    for name in names:
        exact = lower.get(name.lower().strip())
        if exact is not None:
            return exact
    for name in names:
        needle = name.lower().strip()
        for key, val in lower.items():
            if needle in key:
                return val
    return None


def _row_has(row: dict, *names: str) -> bool:
    return any(_row_get(row, name) is not None for name in names)


def _extract_ads_report_sheets(content: bytes) -> dict[str, list[dict]]:
    """Extract usable report rows from either a multi-tab workbook or a single-report workbook."""
    from openpyxl import load_workbook

    extracted = {
        "advertised": [],
        "purchased": [],
        "search_terms": [],
        "targeting": [],
        "placement": [],
    }
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = _worksheet_dicts(ws)
            if not rows:
                continue
            sample = next((row for row in rows if any(v not in (None, "") for v in row.values())), rows[0])
            sheet_key = "".join(ch for ch in ws.title.lower() if ch.isalnum())
            has_advertised_asin = _row_has(sample, "Advertised ASIN")
            has_purchased_asin = _row_has(sample, "Purchased ASIN")
            has_search_term = _row_has(sample, "Customer Search Term", "Search Term")
            has_targeting = _row_has(sample, "Targeting", "Target")
            has_placement = _row_has(sample, "Placement")

            if has_placement or "placement" in sheet_key or sheet_key == "placements":
                extracted["placement"].extend(rows)
            elif has_search_term or ("search" in sheet_key and "term" in sheet_key):
                extracted["search_terms"].extend(rows)
            elif has_advertised_asin and has_purchased_asin:
                extracted["purchased"].extend(rows)
            elif has_targeting or "targeting" in sheet_key:
                extracted["targeting"].extend(rows)
            elif has_advertised_asin or ("advertised" in sheet_key and "product" in sheet_key):
                extracted["advertised"].extend(rows)
    finally:
        wb.close()
    return extracted


def _parse_ads_report_rows(
    advertised: list[dict],
    purchased: list[dict],
    search_terms: list[dict],
    targeting: list[dict],
    placement: list[dict],
    allowed_asins: Optional[set[str]] = None,
    allowed_keywords: Optional[set[str]] = None,
) -> dict[str, list[ReportCampaignRow]]:
    """Return map key 'ASIN|||keyword' -> campaign/ad group rows from combined Amazon ads report rows."""

    asins_by_campaign_adgroup: dict[tuple[str, str], set[str]] = {}
    asins_by_campaign: dict[str, set[str]] = {}
    for row in advertised + purchased:
        asin = _norm_text(_row_get(row, "Advertised ASIN")).upper()
        campaign = _norm_text(_row_get(row, "Campaign Name"))
        ad_group = _norm_text(_row_get(row, "Ad Group Name"))
        if not asin or not campaign:
            continue
        if allowed_asins is not None and asin not in allowed_asins:
            continue
        asins_by_campaign_adgroup.setdefault((_norm_key(campaign), _norm_key(ad_group)), set()).add(asin)
        asins_by_campaign.setdefault(_norm_key(campaign), set()).add(asin)

    placement_top: dict[str, dict] = {}
    for row in placement:
        campaign = _norm_text(_row_get(row, "Campaign Name"))
        placement_name = _norm_key(_row_get(row, "Placement"))
        if not campaign or "top of search" not in placement_name:
            continue
        spend = _num(_row_get(row, "Spend"))
        sales = _num(_row_get(row, "7 Day Total Sales"))
        placement_top[_norm_key(campaign)] = {
            "acos": spend / sales if sales > 0 else None,
            "orders": _num(_row_get(row, "7 Day Total Orders")),
        }

    def asins_for(campaign: str, ad_group: str) -> set[str]:
        asins = asins_by_campaign_adgroup.get((_norm_key(campaign), _norm_key(ad_group)))
        if asins:
            return asins
        return asins_by_campaign.get(_norm_key(campaign), set())

    raw_rows: list[tuple[str, str, ReportCampaignRow]] = []

    for row in search_terms:
        campaign = _norm_text(_row_get(row, "Campaign Name"))
        ad_group = _norm_text(_row_get(row, "Ad Group Name"))
        keyword = _norm_text(_row_get(row, "Customer Search Term", "Search Term", "Targeting"))
        if not keyword:
            continue
        keyword_key = _keyword_key(keyword)
        if allowed_keywords is not None and keyword_key not in allowed_keywords:
            continue
        clicks = _num(_row_get(row, "Clicks"))
        impressions = _num(_row_get(row, "Impressions"))
        spend = _num(_row_get(row, "Spend"))
        sales = _num(_row_get(row, "7 Day Total Sales"))
        orders = _num(_row_get(row, "7 Day Total Orders"))
        top = placement_top.get(_norm_key(campaign), {})
        campaign_row = ReportCampaignRow(
            campaign=campaign or "Unknown Campaign",
            ad_group=ad_group,
            match_type=_norm_text(_row_get(row, "Match Type")) or "AUTO",
            source="Search Term",
            impressions=impressions,
            ctr=clicks / impressions if impressions > 0 else _rate(_row_get(row, "Click-Thru Rate")),
            cpc=spend / clicks if clicks > 0 else _num(_row_get(row, "Cost Per Click")),
            spend=spend,
            sales=sales,
            orders=orders,
            cvr=orders / clicks if clicks > 0 else 0,
            acos=spend / sales if sales > 0 else None,
            placement_top_acos=top.get("acos"),
            placement_top_orders=top.get("orders"),
        )
        for asin in asins_for(campaign, ad_group):
            raw_rows.append((asin, keyword_key, campaign_row))

    for row in targeting:
        campaign = _norm_text(_row_get(row, "Campaign Name"))
        ad_group = _norm_text(_row_get(row, "Ad Group Name"))
        keyword = _norm_text(_row_get(row, "Targeting", "Target", "Keyword"))
        if not keyword:
            continue
        keyword_key = _keyword_key(keyword)
        if allowed_keywords is not None and keyword_key not in allowed_keywords:
            continue
        clicks = _num(_row_get(row, "Clicks"))
        impressions = _num(_row_get(row, "Impressions"))
        spend = _num(_row_get(row, "Spend"))
        sales = _num(_row_get(row, "7 Day Total Sales"))
        orders = _num(_row_get(row, "7 Day Total Orders"))
        top = placement_top.get(_norm_key(campaign), {})
        campaign_row = ReportCampaignRow(
            campaign=campaign or "Unknown Campaign",
            ad_group=ad_group,
            match_type=_norm_text(_row_get(row, "Match Type")) or "AUTO",
            source="Targeting",
            impressions=impressions,
            ctr=clicks / impressions if impressions > 0 else _rate(_row_get(row, "Click-Thru Rate")),
            cpc=spend / clicks if clicks > 0 else _num(_row_get(row, "Cost Per Click")),
            spend=spend,
            sales=sales,
            orders=orders,
            cvr=orders / clicks if clicks > 0 else 0,
            acos=spend / sales if sales > 0 else None,
            top_of_search_is=_rate(_row_get(row, "Top-of-search Impression Share")),
            placement_top_acos=top.get("acos"),
            placement_top_orders=top.get("orders"),
        )
        for asin in asins_for(campaign, ad_group):
            raw_rows.append((asin, keyword_key, campaign_row))

    grouped: dict[str, dict[tuple[str, str, str, str], list[ReportCampaignRow]]] = {}
    for asin, keyword, row in raw_rows:
        map_key = f"{asin}|||{keyword}"
        group_key = (_norm_key(row.campaign), _norm_key(row.ad_group), _norm_key(row.match_type), row.source)
        grouped.setdefault(map_key, {}).setdefault(group_key, []).append(row)

    result: dict[str, list[ReportCampaignRow]] = {}
    for map_key, groups in grouped.items():
        result[map_key] = []
        for rows in groups.values():
            total_impressions = sum(r.impressions for r in rows)
            total_spend = sum(r.spend for r in rows)
            total_sales = sum(r.sales for r in rows)
            total_orders = sum(r.orders for r in rows)
            total_clicks = sum((r.ctr * r.impressions) for r in rows)
            first = rows[0]
            result[map_key].append(ReportCampaignRow(
                campaign=first.campaign,
                ad_group=first.ad_group,
                match_type=first.match_type,
                source=first.source,
                impressions=total_impressions,
                ctr=total_clicks / total_impressions if total_impressions > 0 else 0,
                cpc=total_spend / total_clicks if total_clicks > 0 else 0,
                spend=total_spend,
                sales=total_sales,
                orders=total_orders,
                cvr=total_orders / total_clicks if total_clicks > 0 else 0,
                acos=total_spend / total_sales if total_sales > 0 else None,
                top_of_search_is=next((r.top_of_search_is for r in rows if r.top_of_search_is is not None), None),
                placement_top_acos=next((r.placement_top_acos for r in rows if r.placement_top_acos is not None), None),
                placement_top_orders=next((r.placement_top_orders for r in rows if r.placement_top_orders is not None), None),
            ))
    return result


def _parse_full_ads_workbook(content: bytes) -> dict[str, list[ReportCampaignRow]]:
    """Return map key 'ASIN|||keyword' -> campaign/ad group rows from full Amazon ads workbook."""
    sheets = _extract_ads_report_sheets(content)
    return _parse_ads_report_rows(
        sheets["advertised"],
        sheets["purchased"],
        sheets["search_terms"],
        sheets["targeting"],
        sheets["placement"],
    )


def _merge_ads_maps(maps: list[dict[str, list[ReportCampaignRow]]]) -> dict[str, list[ReportCampaignRow]]:
    merged: dict[str, list[ReportCampaignRow]] = {}
    for ads_map in maps:
        for key, rows in ads_map.items():
            merged.setdefault(key, []).extend(rows)
    return merged


async def _run_report_ai(keywords: list[ReportKeyword], target_acos: float) -> dict:
    batches = [keywords[i:i + REPORT_BATCH_SIZE] for i in range(0, len(keywords), REPORT_BATCH_SIZE)]
    semaphore = asyncio.Semaphore(AI_CONCURRENCY)

    def run_batch(batch: list[ReportKeyword]) -> tuple[list, list]:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        msg = _anthropic_create_message(
            client,
            model="claude-sonnet-4-6",
            max_tokens=7000,
            system=REPORT_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": _make_report_batch_prompt(batch, target_acos)}],
        )
        raw = msg.content[0].text.strip()
        try:
            return _report_json_parts(raw)
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to parse AI response: {e}. Raw: {raw[:300]}"
            )

    async def run_limited(batch: list[ReportKeyword]) -> tuple[list, list]:
        async with semaphore:
            return await asyncio.to_thread(run_batch, batch)

    parts = await asyncio.gather(*[run_limited(batch) for batch in batches])
    all_rows = [row for rows, _ in parts for row in rows]
    all_summaries = [summary for _, summaries in parts for summary in summaries]
    return {"rows": all_rows, "summaries": all_summaries}


@app.post("/api/report-analysis")
async def report_analysis(req: ReportAnalysisRequest):
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured.")
    if not req.keywords:
        raise HTTPException(status_code=400, detail="No keywords provided.")
    if len(req.keywords) > 50:
        raise HTTPException(status_code=400, detail="Maximum 50 ASIN+keyword groups per analysis.")

    return await _run_report_ai(req.keywords, req.target_acos)


@app.post("/api/full-report-analysis")
async def full_report_analysis(
    filter_file: UploadFile = File(...),
    ads_files: Optional[list[UploadFile]] = File(default=None),
    ads_file: Optional[UploadFile] = File(default=None),
    target_acos: float = Form(20.0),
):
    if not SERPAPI_KEY:
        raise HTTPException(status_code=500, detail="SERPAPI_KEY not configured.")
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured.")

    filter_name = filter_file.filename or ""
    all_ads_files = list(ads_files or [])
    if ads_file is not None:
        all_ads_files.append(ads_file)

    if not (filter_name.lower().endswith(".csv") or filter_name.lower().endswith((".xlsx", ".xlsm"))):
        raise HTTPException(status_code=400, detail="ASIN + Search Terms file must be .csv, .xlsx, or .xlsm.")
    if not all_ads_files:
        raise HTTPException(status_code=400, detail="Upload at least one full Amazon Ads .xlsx or .xlsm workbook.")
    for uploaded in all_ads_files:
        ads_name = uploaded.filename or ""
        if not ads_name.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail=f"{ads_name or 'Ads report'} must be an .xlsx or .xlsm workbook.")

    entries = parse_upload(await filter_file.read(), filter_name)
    if not entries:
        raise HTTPException(status_code=400, detail="No valid ASIN + search term rows found in filter file.")

    combined_sheets = {
        "advertised": [],
        "purchased": [],
        "search_terms": [],
        "targeting": [],
        "placement": [],
    }
    parsed_file_count = 0
    parse_errors = []
    for uploaded in all_ads_files:
        ads_name = uploaded.filename or "ads workbook"
        try:
            sheets = _extract_ads_report_sheets(await uploaded.read())
            if any(sheets.values()):
                parsed_file_count += 1
                for key, rows in sheets.items():
                    combined_sheets[key].extend(rows)
            else:
                parse_errors.append(f"{ads_name}: no recognized Amazon Ads report columns found")
        except Exception as e:
            parse_errors.append(f"{ads_name}: {e}")

    requested_asins = {entry["asin"] for entry in entries}
    requested_keywords = {
        _keyword_key(term)
        for entry in entries
        for term in entry["terms"]
    }
    ads_map = _parse_ads_report_rows(
        combined_sheets["advertised"],
        combined_sheets["purchased"],
        combined_sheets["search_terms"],
        combined_sheets["targeting"],
        combined_sheets["placement"],
        allowed_asins=requested_asins,
        allowed_keywords=requested_keywords,
    )
    if not ads_map:
        details = "; ".join(parse_errors[:5]) if parse_errors else "No usable ad rows found."
        sheet_counts = ", ".join(f"{key}={len(rows)}" for key, rows in combined_sheets.items())
        raise HTTPException(status_code=400, detail=f"Could not parse usable campaign/search term data from uploaded ads workbook(s). {details} Parsed rows: {sheet_counts}.")

    keyword_groups = []
    found_rank_count = 0
    keyword_only_matches = 0
    asin_only_matches = 0
    ads_asins = {key.split("|||", 1)[0] for key in ads_map}
    ads_keywords = {key.split("|||", 1)[1] for key in ads_map if "|||" in key}
    rank_results = await _rank_entries(entries)
    for entry_result in rank_results:
        asin = entry_result["asin"]
        for ranking in entry_result["rankings"]:
            if not ranking.get("found"):
                continue
            found_rank_count += 1
            key = f"{asin}|||{_keyword_key(ranking.get('term'))}"
            campaigns = ads_map.get(key, [])
            if campaigns:
                keyword_groups.append(ReportKeyword(
                    asin=asin,
                    keyword=ranking["term"],
                    organic_rank=ranking["position"],
                    campaigns=campaigns[:12],
                ))
            else:
                term_key = _keyword_key(ranking.get("term"))
                if term_key in ads_keywords:
                    keyword_only_matches += 1
                if asin in ads_asins:
                    asin_only_matches += 1

    if not keyword_groups:
        reasons = []
        if found_rank_count == 0:
            reasons.append("No uploaded keywords were found organically in the checked Amazon result pages.")
        if not combined_sheets["advertised"]:
            reasons.append("No Advertised Product rows were uploaded, so campaign/ad group to ASIN mapping may be incomplete. Upload the Sponsored Products Advertised Product report or the full multi-tab Sponsored Products workbook.")
        if keyword_only_matches > 0 and asin_only_matches == 0:
            reasons.append("Some keywords exist in the ads report, but the ASINs from your ranking file do not match the ASINs mapped in the ads reports.")
        elif asin_only_matches > 0 and keyword_only_matches == 0:
            reasons.append("Some ASINs exist in the ads report, but the exact keywords/search terms from your ranking file do not match the ads search terms.")
        elif keyword_only_matches > 0 and asin_only_matches > 0:
            reasons.append("The ads reports contain some matching ASINs and some matching keywords, but not the same ASIN + keyword pairs together.")
        if not reasons:
            reasons.append("Ads data was parsed, but no exact ASIN + keyword pair matched the organic ranking results.")
        return {
            "rows": [],
            "summaries": [],
            "rank_results": rank_results,
            "matched_groups": 0,
            "ads_files_parsed": parsed_file_count,
            "message": "Organic rankings ran, but no matching ASIN+keyword campaign data was found. " + " ".join(reasons),
        }

    analysis = await _run_report_ai(keyword_groups[:50], target_acos)
    analysis.update({
        "rank_results": rank_results,
        "matched_groups": len(keyword_groups),
        "ads_files_parsed": parsed_file_count,
    })
    return analysis


def _first_worksheet_rows(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = _worksheet_dicts(ws)
            if rows:
                return rows
        return []
    finally:
        wb.close()


def _parse_seo_ads_rows(targeting_content: bytes, advertised_content: bytes) -> list[dict]:
    advertised_rows = _first_worksheet_rows(advertised_content)
    targeting_rows = _first_worksheet_rows(targeting_content)

    asin_map: dict[tuple[str, str], set[str]] = {}
    for row in advertised_rows:
        campaign = _norm_key(_row_get(row, "Campaign Name", "Campaign"))
        ad_group = _norm_key(_row_get(row, "Ad Group Name", "Ad Group"))
        asin = _norm_text(_row_get(row, "Advertised ASIN")).upper()
        if campaign and asin:
            asin_map.setdefault((campaign, ad_group), set()).add(asin)

    aggregated: dict[tuple[str, str, str], dict] = {}
    valid_match_types = {"BROAD", "PHRASE", "EXACT"}
    for row in targeting_rows:
        match_type = _norm_text(_row_get(row, "Match Type")).upper()
        keyword = _norm_text(_row_get(row, "Targeting", "Target", "Keyword"))
        orders = _num(_row_get(row, "7 Day Total Orders (#)", "7 Day Total Orders", "Orders"))
        if match_type not in valid_match_types or orders <= 0:
            continue
        if not keyword or keyword in {"-", "*"} or "placeholder keyword" in keyword.lower():
            continue

        campaign = _norm_key(_row_get(row, "Campaign Name", "Campaign"))
        ad_group = _norm_key(_row_get(row, "Ad Group Name", "Ad Group"))
        asins = asin_map.get((campaign, ad_group), set())
        if not asins:
            continue

        impressions = _num(_row_get(row, "Impressions"))
        clicks = _num(_row_get(row, "Clicks"))
        spend = _num(_row_get(row, "Spend"))
        sales = _num(_row_get(row, "7 Day Total Sales", "Sales"))
        keyword_key = _keyword_key(keyword)

        for asin in asins:
            key = (asin, keyword_key, match_type)
            item = aggregated.setdefault(key, {
                "asin": asin,
                "keyword": keyword.strip(),
                "keyword_key": keyword_key,
                "match_type": match_type,
                "impressions": 0.0,
                "clicks": 0.0,
                "spend": 0.0,
                "orders": 0.0,
                "sales": 0.0,
            })
            item["impressions"] += impressions
            item["clicks"] += clicks
            item["spend"] += spend
            item["orders"] += orders
            item["sales"] += sales

    return list(aggregated.values())


def _marketplace_domain(marketplace: str) -> str:
    mapping = {
        "amazon.com": "US",
        "amazon.co.uk": "UK",
        "amazon.de": "DE",
        "amazon.fr": "FR",
        "amazon.it": "IT",
        "amazon.es": "ES",
        "amazon.ca": "CA",
        "amazon.com.mx": "MX",
        "amazon.com.br": "BR",
        "amazon.in": "IN",
        "amazon.co.jp": "JP",
        "amazon.com.au": "AU",
    }
    value = marketplace.strip().lower().replace("www.", "")
    return mapping.get(value, "US")


def _canopy_results(data) -> list[dict]:
    if isinstance(data, dict):
        product_results = data.get("productResults")
        if isinstance(product_results, dict):
            results = product_results.get("results")
            if isinstance(results, list):
                return [item for item in results if isinstance(item, dict)]

        for key in ("amazonProductSearchResults", "searchResults", "search_results"):
            value = data.get(key)
            if isinstance(value, dict):
                found = _canopy_results(value)
                if found:
                    return found

        for key in ("products", "items", "results"):
            value = data.get(key)
            if isinstance(value, list):
                products = [
                    item for item in value
                    if isinstance(item, dict) and (
                        item.get("asin") or item.get("ASIN") or item.get("productAsin")
                    )
                ]
                if products:
                    return products
        for value in data.values():
            if isinstance(value, dict):
                found = _canopy_results(value)
                if found:
                    return found
    return []


def _canopy_is_sponsored(item: dict) -> bool:
    for key in ("sponsored", "isSponsored", "is_sponsored", "sponsoredResult"):
        value = item.get(key)
        if isinstance(value, bool):
            return value
        if isinstance(value, str) and value.lower() in {"true", "yes", "sponsored"}:
            return True
    return False


async def _serpapi_backup_rank(
    client: httpx.AsyncClient,
    asin: str,
    term: str,
    marketplace: str,
    canopy_error: str,
) -> dict:
    if not SERPAPI_KEY:
        raise HTTPException(
            status_code=502,
            detail=f"{canopy_error} SerpAPI backup is not configured.",
        )
    try:
        result = await search_term_ranking(client, asin, term, marketplace)
    except HTTPException as exc:
        raise HTTPException(
            status_code=502,
            detail=f"{canopy_error} SerpAPI backup also failed: {exc.detail}",
        ) from exc
    return {
        "asin": asin.upper(),
        "marketplace": marketplace,
        **result,
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "provider": "SerpAPI backup",
    }


async def _canopy_rank(
    client: httpx.AsyncClient,
    asin: str,
    term: str,
    marketplace: str,
) -> dict:
    asin = asin.upper()
    checked_at = datetime.now(timezone.utc).isoformat()
    if not CANOPY_API_KEY:
        return await _serpapi_backup_rank(
            client, asin, term, marketplace,
            "Canopy API is not configured.",
        )

    for page in range(1, CANOPY_MAX_PAGES + 1):
        try:
            response = await client.get(
                CANOPY_SEARCH_URL,
                params={
                    "searchTerm": term,
                    "domain": _marketplace_domain(marketplace),
                    "page": page,
                },
                headers={
                    "API-KEY": CANOPY_API_KEY,
                    "Authorization": f"Bearer {CANOPY_API_KEY}",
                },
                timeout=45,
            )
        except Exception as exc:
            return await _serpapi_backup_rank(
                client, asin, term, marketplace,
                f"Canopy API request failed: {exc}.",
            )

        if response.status_code != 200:
            return await _serpapi_backup_rank(
                client, asin, term, marketplace,
                f"Canopy API error: {response.text[:300] or f'HTTP {response.status_code}'}.",
            )
        try:
            data = response.json()
        except Exception:
            return await _serpapi_backup_rank(
                client, asin, term, marketplace,
                f"Canopy API returned invalid JSON: {response.text[:250]}.",
            )

        results = _canopy_results(data)
        for index, item in enumerate(results, 1):
            if _canopy_is_sponsored(item):
                continue
            item_asin = _norm_text(
                item.get("asin") or item.get("ASIN") or item.get("productAsin")
            ).upper()
            if item_asin == asin:
                raw_position = int(_num(item.get("position")) or index)
                rank = raw_position if raw_position > 20 else (page - 1) * 20 + raw_position
                return {
                    "asin": asin,
                    "marketplace": marketplace,
                    "term": term,
                    "position": rank,
                    "page": page,
                    "found": True,
                    "checked_at": checked_at,
                    "provider": "Canopy",
                }
        if not results:
            break

    return {
        "asin": asin,
        "marketplace": marketplace,
        "term": term,
        "position": None,
        "page": None,
        "found": False,
        "checked_at": checked_at,
        "provider": "Canopy",
    }


async def _canopy_rank_entries(entries: list[dict]) -> list[dict]:
    jobs = [
        (entry["asin"], entry["marketplace"], term)
        for entry in entries
        for term in entry["terms"]
    ]
    limits = httpx.Limits(
        max_connections=max(CANOPY_CONCURRENCY * 2, 10),
        max_keepalive_connections=max(CANOPY_CONCURRENCY, 5),
    )
    semaphore = asyncio.Semaphore(CANOPY_CONCURRENCY)

    async def run_limited(asin: str, marketplace: str, term: str) -> dict:
        async with semaphore:
            return await _canopy_rank(client, asin, term, marketplace)

    async with httpx.AsyncClient(limits=limits, timeout=50) as client:
        return await asyncio.gather(*[
            run_limited(asin, marketplace, term)
            for asin, marketplace, term in jobs
        ])


def _write_report_files(prefix: str, headers: list[str], rows: list[list]) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    export_dir = Path(__file__).parent / "static" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    token = uuid4().hex[:8]
    csv_name = f"{prefix}_{token}.csv"
    xlsx_name = f"{prefix}_{token}.xlsx"

    with (export_dir / csv_name).open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = prefix[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="FF9900")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        values = [str(cell.value or "") for cell in column[:200]]
        ws.column_dimensions[column[0].column_letter].width = min(max(map(len, values), default=10) + 2, 45)
    wb.save(export_dir / xlsx_name)

    return {
        "csv_url": f"/static/exports/{csv_name}",
        "xlsx_url": f"/static/exports/{xlsx_name}",
        "csv_name": csv_name,
        "xlsx_name": xlsx_name,
    }


@app.post("/api/seo-gap-analysis")
async def seo_gap_analysis(
    targeting_file: UploadFile = File(...),
    advertised_file: UploadFile = File(...),
    rank_file: UploadFile = File(...),
):
    if not CANOPY_API_KEY and not SERPAPI_KEY:
        raise HTTPException(
            status_code=500,
            detail="Neither CANOPY_API_KEY nor SERPAPI_KEY is configured.",
        )

    target_name = targeting_file.filename or ""
    advertised_name = advertised_file.filename or ""
    rank_name = rank_file.filename or ""
    if not target_name.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Targeting Report must be XLSX or XLSM.")
    if not advertised_name.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Advertised Products Report must be XLSX or XLSM.")
    if not rank_name.lower().endswith((".csv", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="ASIN Rank Template must be CSV, XLSX, or XLSM.")

    entries = parse_upload(await rank_file.read(), rank_name)
    if not entries:
        raise HTTPException(status_code=400, detail="No valid ASIN + search terms found in the rank template.")

    ads_rows = _parse_seo_ads_rows(
        await targeting_file.read(),
        await advertised_file.read(),
    )
    ranks = await _canopy_rank_entries(entries)

    rank_headers = [
        "ASIN", "Marketplace", "Search Term", "Organic Rank", "Rank Page", "Checked At"
    ]
    rank_rows = [
        [
            item["asin"],
            item["marketplace"],
            item["term"],
            item["position"] if item["found"] else "Not ranked",
            item["page"] if item["found"] else "",
            item["checked_at"],
        ]
        for item in ranks
    ]

    rank_lookup = {
        (item["asin"], _keyword_key(item["term"])): item
        for item in ranks
    }
    marketplace_by_asin = {entry["asin"]: entry["marketplace"] for entry in entries}

    ads_lookup: dict[tuple[str, str], dict] = {}
    for row in ads_rows:
        key = (row["asin"], row["keyword_key"])
        item = ads_lookup.setdefault(key, {
            **row,
            "match_types": set(),
            "impressions": 0.0,
            "clicks": 0.0,
            "spend": 0.0,
            "orders": 0.0,
            "sales": 0.0,
        })
        item["match_types"].add(row["match_type"])
        for metric in ("impressions", "clicks", "spend", "orders", "sales"):
            item[metric] += row[metric]

    keyword_labels = {
        (item["asin"], _keyword_key(item["term"])): item["term"] for item in ranks
    }
    for key, item in ads_lookup.items():
        keyword_labels.setdefault(key, item["keyword"])

    all_keys = set(keyword_labels)
    gap_headers = [
        "ASIN", "Keyword", "Match Type", "In Ads", "In Organic", "Organic Rank",
        "Bucket", "Impressions", "Clicks", "CTR", "CVR", "Orders", "Sales",
        "Spend", "ROAS", "Priority Score"
    ]
    gap_rows = []
    bucket_counts = {"Winning": 0, "Paid Only": 0, "Organic Only": 0, "Gap": 0}
    for asin, keyword_key in sorted(all_keys):
        ad = ads_lookup.get((asin, keyword_key))
        rank = rank_lookup.get((asin, keyword_key))
        in_ads = ad is not None
        in_organic = bool(rank and rank["found"])
        if in_ads and in_organic:
            bucket = "Winning"
        elif in_ads:
            bucket = "Paid Only"
        elif in_organic:
            bucket = "Organic Only"
        else:
            bucket = "Gap"
        bucket_counts[bucket] += 1

        impressions = ad["impressions"] if ad else None
        clicks = ad["clicks"] if ad else None
        spend = ad["spend"] if ad else None
        orders = ad["orders"] if ad else None
        sales = ad["sales"] if ad else None
        ctr = (clicks / impressions * 100) if ad and impressions > 0 else None
        cvr = (orders / clicks * 100) if ad and clicks > 0 else None
        roas = (sales / spend) if ad and spend > 0 else None
        priority = (
            cvr * roas * math.log(impressions + 1)
            if bucket == "Paid Only" and cvr is not None and roas is not None
            else None
        )

        gap_rows.append([
            asin,
            keyword_labels[(asin, keyword_key)],
            ", ".join(sorted(ad["match_types"])) if ad else "",
            "YES" if in_ads else "NO",
            "YES" if in_organic else "NO",
            rank["position"] if in_organic else "Not ranked",
            bucket,
            round(impressions, 2) if impressions is not None else "",
            round(clicks, 2) if clicks is not None else "",
            round(ctr, 4) if ctr is not None else "",
            round(cvr, 4) if cvr is not None else "",
            round(orders, 2) if orders is not None else "",
            round(sales, 2) if sales is not None else "",
            round(spend, 2) if spend is not None else "",
            round(roas, 4) if roas is not None else "",
            round(priority, 4) if priority is not None else "",
        ])

    rank_files = _write_report_files("rank_report", rank_headers, rank_rows)
    gap_files = _write_report_files("seo_gap_report", gap_headers, gap_rows)

    return {
        "rank_report": rank_files,
        "seo_gap_report": gap_files,
        "summary": {
            "asins": len({entry["asin"] for entry in entries}),
            "rank_checks": len(ranks),
            "ranked": sum(1 for item in ranks if item["found"]),
            "ads_keywords": len(ads_lookup),
            "buckets": bucket_counts,
            "providers": {
                "canopy": sum(1 for item in ranks if item.get("provider") == "Canopy"),
                "serpapi_backup": sum(1 for item in ranks if item.get("provider") == "SerpAPI backup"),
            },
        },
    }


@app.post("/api/save-csv")
async def save_csv(req: CSVExportRequest):
    safe_name = Path(req.filename).name
    if not safe_name.lower().endswith(".csv"):
        safe_name += ".csv"

    export_dir = Path(__file__).parent / "static" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    output_path = export_dir / safe_name

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(req.rows)

    return {
        "filename": safe_name,
        "url": f"/static/exports/{safe_name}",
        "path": str(output_path),
    }


app.mount("/static", StaticFiles(directory="static"), name="static")
