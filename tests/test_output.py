"""Tests for catalog_engine.output.write_outputs.

Uses the real sample Category Listings Report (skipped if absent) so the
upload.xlsx header-copy and column location are exercised against a genuine
Amazon template, with fabricated GeneratedContent on ~3 records.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from catalog_engine.models import (
    ComplianceLogEntry,
    GeneratedContent,
    Issue,
    Severity,
    SkuOutcome,
)
from catalog_engine.output import write_outputs
from catalog_engine.parser import parse_flat_file

SAMPLE = Path("/Users/denizolmez/Downloads/Category+Listings+Report_07-02-2026.xlsm")

# openpyxl emits a harmless "Data Validation extension" UserWarning on load.
pytestmark = pytest.mark.filterwarnings("ignore:Data Validation extension")

RUN_ID = "run-test-0001"
SCRIPT_INJECTION = "<script>alert('xss')</script> sturdy build"


@pytest.fixture(scope="module")
def parse():
    if not SAMPLE.exists():
        pytest.skip(f"sample file not available: {SAMPLE}")
    return parse_flat_file(SAMPLE, seller_id="seller-test")


def _gen(sku: str, n: int) -> GeneratedContent:
    return GeneratedContent(
        sku=sku,
        title=f"Acme Widget {n} Oak 11x17 Frame",
        item_highlight=f"Highlight for {sku}",
        bullets=[
            f"DURABLE BUILD - bullet one for product {n}",
            f"EASY SETUP - bullet two for product {n}",
            f"VERSATILE USE - bullet three for product {n}",
        ],
        description=f"A generated description for product {n}. Plain text only.",
        search_terms=f"widget frame oak decor gift {n}",
        model="claude-sonnet-5",
    )


@pytest.fixture(scope="module")
def outcomes(parse):
    r1, r2, r3 = parse.records[:3]
    ok1 = SkuOutcome(record=r1, generated=_gen(r1.sku, 1), status="ok")
    ok2 = SkuOutcome(record=r2, generated=_gen(r2.sku, 2), status="ok")

    gen3 = _gen(r3.sku, 3)
    gen3.bullets[1] = SCRIPT_INJECTION  # must be escaped in report.html
    review = SkuOutcome(
        record=r3,
        generated=gen3,
        status="needs_review",
        issues=[
            Issue(
                sku=r3.sku,
                field_name="bullet_2",
                rule="unsupported_claim",
                severity=Severity.ERROR,
                message="claim not supported by source attributes",
            ),
            Issue(
                sku=r3.sku,
                field_name="input:material",
                rule="missing_mandatory_attribute",
                severity=Severity.WARNING,
                message="required attribute 'material' is empty in the source file",
            ),
        ],
        compliance_log=[
            ComplianceLogEntry(
                seller_id="seller-test",
                run_id=RUN_ID,
                sku=r3.sku,
                field_name="bullet_2",
                category="hallucination",
                removed_text="",
                reason="flagged unsupported claim in bullet 2",
            )
        ],
    )
    return [ok1, ok2, review]


@pytest.fixture(scope="module")
def artifacts(parse, outcomes, tmp_path_factory):
    out_dir = tmp_path_factory.mktemp("outputs")
    return write_outputs(parse, outcomes, RUN_ID, out_dir)


@pytest.fixture(scope="module")
def upload_ws(artifacts):
    wb = openpyxl.load_workbook(artifacts["upload.xlsx"])
    yield wb["Template"]
    wb.close()


def _source_header_rows(data_row: int):
    wb = openpyxl.load_workbook(SAMPLE, read_only=True, data_only=True)
    try:
        rows = list(
            wb["Template"].iter_rows(min_row=1, max_row=data_row - 1,
                                     values_only=True)
        )
    finally:
        wb.close()
    return rows


def _col(parse, base: str, instance: int = 1):
    for c in parse.columns:
        if c.base == base and c.instance == instance and c.leaf in ("value", ""):
            return c
    raise AssertionError(f"column {base}#{instance} not found in parse.columns")


def test_returns_all_artifacts(artifacts):
    assert set(artifacts) == {
        "upload.xlsx", "results.json", "results.csv", "results.xlsx",
        "report.html", "input_issues.json",
    }
    for path in artifacts.values():
        assert Path(path).exists()


def test_upload_a1_settings_identical(parse, upload_ws):
    src_rows = _source_header_rows(parse.settings.data_row)
    assert upload_ws.cell(1, 1).value == src_rows[0][0]
    assert str(upload_ws.cell(1, 1).value).startswith("settings=")


def test_upload_header_rows_match_source(parse, upload_ws):
    src_rows = _source_header_rows(parse.settings.data_row)
    for r, src_row in enumerate(src_rows, 1):
        for c in range(1, 41):  # first 40 columns
            src_val = src_row[c - 1] if c - 1 < len(src_row) else None
            assert upload_ws.cell(r, c).value == src_val, (
                f"header mismatch at row {r} col {c}"
            )


def test_only_ok_skus_written_as_data_rows(parse, outcomes, upload_ws):
    sku_col = _col(parse, "contribution_sku")
    written = []
    for row in range(parse.settings.data_row, upload_ws.max_row + 1):
        val = upload_ws.cell(row, sku_col.index + 1).value
        if val:
            written.append(str(val))
    ok_skus = [o.record.sku for o in outcomes if o.status == "ok"]
    assert written == ok_skus  # needs_review SKU must NOT appear
    assert outcomes[2].record.sku not in written


def test_data_row_has_only_expected_cells(parse, outcomes, upload_ws):
    settings = parse.settings
    row = settings.data_row  # first ok outcome
    outcome = outcomes[0]
    gen = outcome.generated

    expected = {
        _col(parse, "contribution_sku").index: outcome.record.sku,
        _col(parse, "product_type").index: outcome.record.product_type,
        _col(parse, "::record_action").index: "partial_update",
        _col(parse, "item_name").index: gen.title,
        _col(parse, "bullet_point", 1).index: gen.bullets[0],
        _col(parse, "bullet_point", 2).index: gen.bullets[1],
        _col(parse, "bullet_point", 3).index: gen.bullets[2],
        _col(parse, "product_description").index: gen.description,
        _col(parse, "generic_keyword", 1).index: gen.search_terms,
    }
    highlight_cols = [c for c in parse.columns if c.base == "title_differentiation"]
    if highlight_cols:
        expected[highlight_cols[0].index] = gen.item_highlight
    max_col = max(c.index for c in parse.columns) + 1
    for idx in range(max_col):
        value = upload_ws.cell(row, idx + 1).value
        if idx in expected:
            assert value == expected[idx], f"col index {idx}"
        else:
            assert value in (None, ""), (
                f"unexpected value {value!r} at column index {idx}"
            )


def test_bullets_4_and_5_left_empty(parse, upload_ws):
    row = parse.settings.data_row
    for instance in (4, 5):
        col = _col(parse, "bullet_point", instance)
        assert upload_ws.cell(row, col.index + 1).value in (None, "")


def test_generic_keyword_2_plus_empty(parse, upload_ws):
    row = parse.settings.data_row
    cols = [c for c in parse.columns
            if c.base == "generic_keyword" and c.instance >= 2
            and c.leaf in ("value", "")]
    for col in cols:
        assert upload_ws.cell(row, col.index + 1).value in (None, "")


def test_record_action_is_partial_update(parse, upload_ws):
    action_col = _col(parse, "::record_action")
    for offset in range(2):  # both ok rows
        row = parse.settings.data_row + offset
        assert upload_ws.cell(row, action_col.index + 1).value == "partial_update"


def test_report_html_contains_skus_and_escapes_injection(artifacts, outcomes):
    html_text = Path(artifacts["report.html"]).read_text(encoding="utf-8")
    for outcome in outcomes:
        assert outcome.record.sku in html_text or (
            # SKUs containing HTML-special chars appear escaped
            __import__("html").escape(outcome.record.sku) in html_text
        )
    assert "<script>" not in html_text  # injection must not survive raw
    assert "&lt;script&gt;" in html_text
    assert "not uploaded — paste into Seller Central manually" in html_text
    assert RUN_ID in html_text


def test_report_orders_needs_review_first(artifacts, outcomes):
    html_text = Path(artifacts["report.html"]).read_text(encoding="utf-8")
    review_sku = outcomes[2].record.sku
    ok_positions = [
        html_text.index(o.record.sku) for o in outcomes if o.status == "ok"
    ]
    assert html_text.index(review_sku) < min(ok_positions)


def test_results_json_structure(artifacts, outcomes):
    data = json.loads(Path(artifacts["results.json"]).read_text(encoding="utf-8"))
    assert data["run_id"] == RUN_ID
    assert data["source_file"].endswith(".xlsm")
    assert "generated_at" in data
    assert len(data["skus"]) == 3

    by_sku = {s["sku"]: s for s in data["skus"]}
    ok = by_sku[outcomes[0].record.sku]
    assert ok["status"] == "ok"
    gen = ok["generated"]
    assert gen["item_highlight"] == outcomes[0].generated.item_highlight
    counts = gen["counts"]
    assert counts["title_chars"] == len(outcomes[0].generated.title)
    assert counts["bullet_chars"] == [
        len(b) for b in outcomes[0].generated.bullets
    ]
    assert counts["search_terms_bytes"] == len(
        outcomes[0].generated.search_terms.encode("utf-8")
    )

    review = by_sku[outcomes[2].record.sku]
    assert review["status"] == "needs_review"
    assert review["issues"][0]["severity"] == "error"
    assert review["compliance_log"][0]["category"] == "hallucination"


def test_input_issues_json(artifacts, outcomes):
    issues = json.loads(
        Path(artifacts["input_issues.json"]).read_text(encoding="utf-8")
    )
    assert len(issues) == 1
    assert issues[0]["field_name"] == "input:material"
    assert issues[0]["sku"] == outcomes[2].record.sku
    assert issues[0]["severity"] == "warning"
