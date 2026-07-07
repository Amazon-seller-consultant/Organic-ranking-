"""Run artifacts: Amazon-uploadable flat file + JSON/HTML review outputs.

`upload.xlsx` is the file a seller uploads to Seller Central, so correctness
beats elegance here:
- header rows (1 .. dataRow-1) are copied VERBATIM from the original source
  workbook — the settings string in Template!A1 must survive unchanged;
- data cells are located through `parse.columns` (base / instance / leaf from
  the template's own attribute row), never by hardcoded index, because column
  positions shift between exports;
- only SKUs with status "ok" are written, as partial_update rows that touch
  ONLY the regenerated fields (blank cells are ignored by Amazon on partial
  update, so the seller's existing bullets 4-5 and every other attribute
  survive untouched).

Nothing in this module publishes anywhere; it only writes files to out_dir.
"""

from __future__ import annotations

import csv
import html
import json
from collections import Counter
from dataclasses import asdict
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any, Optional

import openpyxl

from .exceptions import CatalogEngineError, FlatFileError
from .models import (
    BULLET_MAX_CHARS,
    DESCRIPTION_MAX_CHARS,
    HIGHLIGHT_MAX_CHARS,
    SEARCH_TERMS_MAX_BYTES,
    TITLE_MAX_CHARS,
    ColumnInfo,
    ParseResult,
    Severity,
    SkuOutcome,
)

TEMPLATE_SHEET = "Template"
RECORD_ACTION_VALUE = "partial_update"

# Display order for per-SKU cards in report.html.
_STATUS_ORDER = {"needs_review": 0, "ok": 1}


# ---------------------------------------------------------------------------
# Column location (never by hardcoded index)
# ---------------------------------------------------------------------------
def _find_column(
    columns: list[ColumnInfo], base: str, instance: int = 1
) -> Optional[ColumnInfo]:
    """Find the data column for `base` #`instance` (the '.value' / bare leaf)."""
    for col in columns:
        if col.base == base and col.instance == instance and col.leaf in ("value", ""):
            return col
    return None


def _require_column(
    columns: list[ColumnInfo], base: str, source_file: str, instance: int = 1
) -> ColumnInfo:
    col = _find_column(columns, base, instance)
    if col is None:
        raise FlatFileError(
            f"cannot build upload.xlsx: column '{base}'"
            + (f" (instance #{instance})" if instance != 1 else "")
            + f" was not found on the attribute row of '{Path(source_file).name}'.",
            "The template is missing a column this tool needs to write. "
            "Re-download the Category Listings Report from Seller Central; if the "
            "column is genuinely absent, this template version is not supported.",
        )
    return col


# ---------------------------------------------------------------------------
# Counting helpers (display only — enforcement lives in validator.py)
# ---------------------------------------------------------------------------
def _chars(s: str) -> int:
    return len(s or "")


def _bytes_utf8(s: str) -> int:
    return len((s or "").encode("utf-8"))


# ---------------------------------------------------------------------------
# upload.xlsx
# ---------------------------------------------------------------------------
def _write_upload_xlsx(
    parse: ParseResult, ok_outcomes: list[SkuOutcome], path: Path
) -> None:
    settings = parse.settings
    cols = parse.columns
    src = parse.source_file

    sku_col = _require_column(cols, "contribution_sku", src)
    pt_col = _require_column(cols, "product_type", src)
    action_col = _require_column(cols, "::record_action", src)
    title_col = _require_column(cols, "item_name", src)
    bullet_cols = [
        _require_column(cols, "bullet_point", src, instance=i) for i in (1, 2, 3)
    ]
    desc_col = _require_column(cols, "product_description", src)
    keyword_col = _require_column(cols, "generic_keyword", src)

    src_path = Path(parse.source_file)
    if not src_path.exists():
        raise FlatFileError(
            f"cannot build upload.xlsx: the original source file '{src_path}' is no "
            "longer available, so the template header rows cannot be copied.",
            "Re-upload the Category Listings Report and re-run.",
        )
    src_wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    try:
        if TEMPLATE_SHEET not in src_wb.sheetnames:
            raise FlatFileError(
                f"cannot build upload.xlsx: '{src_path.name}' has no "
                f"'{TEMPLATE_SHEET}' sheet."
            )
        src_ws = src_wb[TEMPLATE_SHEET]
        header_rows = list(
            src_ws.iter_rows(
                min_row=1, max_row=settings.data_row - 1, values_only=True
            )
        )
    finally:
        src_wb.close()

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = TEMPLATE_SHEET

    # Rows 1 .. data_row-1 verbatim (every column, plain values). This keeps
    # the A1 settings string byte-identical and preserves the blank row(s)
    # between the attribute row and the data row.
    for r, row in enumerate(header_rows, 1):
        for c, value in enumerate(row, 1):
            if value is not None:
                ws.cell(row=r, column=c, value=value)

    row_no = settings.data_row
    for outcome in ok_outcomes:
        gen = outcome.generated
        if gen is None:
            raise CatalogEngineError(
                f"SKU {outcome.record.sku}: status is 'ok' but no generated content "
                "was provided — refusing to write an empty upload row."
            )

        def put(col: ColumnInfo, value: str) -> None:
            if value:
                ws.cell(row=row_no, column=col.index + 1, value=value)

        put(sku_col, outcome.record.sku)
        put(pt_col, outcome.record.product_type)
        put(action_col, RECORD_ACTION_VALUE)
        put(title_col, gen.title)
        # Bullets 1-3 only. #4/#5 stay EMPTY on purpose: partial update ignores
        # blank cells, so the seller's existing bullets 4-5 are preserved.
        for col, bullet in zip(bullet_cols, gen.bullets[:3]):
            put(col, bullet)
        put(desc_col, gen.description)
        # Full space-separated search-terms string in generic_keyword #1 only.
        put(keyword_col, gen.search_terms)
        row_no += 1

    out_wb.save(path)


# ---------------------------------------------------------------------------
# results.json / input_issues.json
# ---------------------------------------------------------------------------
def _json_default(obj: Any) -> Any:
    if isinstance(obj, Enum):
        return obj.value
    return str(obj)


def _dump_json(data: Any, path: Path) -> None:
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, default=_json_default),
        encoding="utf-8",
    )


def _generated_dict(outcome: SkuOutcome) -> Optional[dict[str, Any]]:
    gen = outcome.generated
    if gen is None:
        return None
    return {
        "title": gen.title,
        "item_highlights": list(gen.item_highlights),
        "bullets": list(gen.bullets),
        "description": gen.description,
        "search_terms": gen.search_terms,
        "counts": {
            "title_chars": _chars(gen.title),
            "bullet_chars": [_chars(b) for b in gen.bullets],
            "description_chars": _chars(gen.description),
            "search_terms_bytes": _bytes_utf8(gen.search_terms),
        },
    }


def _write_results_json(
    parse: ParseResult, outcomes: list[SkuOutcome], run_id: str, path: Path
) -> None:
    data = {
        "run_id": run_id,
        "source_file": parse.source_file,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "skus": [
            {
                "sku": o.record.sku,
                "status": o.status,
                "skip_reason": o.skip_reason,
                "product_type": o.record.product_type,
                "generated": _generated_dict(o),
                "issues": [asdict(i) for i in o.issues],
                "compliance_log": [asdict(e) for e in o.compliance_log],
            }
            for o in outcomes
        ],
    }
    _dump_json(data, path)


def _write_input_issues_json(outcomes: list[SkuOutcome], path: Path) -> None:
    issues = [
        asdict(issue)
        for outcome in outcomes
        for issue in outcome.issues
        if issue.field_name.startswith("input:")
    ]
    _dump_json(issues, path)


# ---------------------------------------------------------------------------
# Subset/merged upload files (approve-flagged flow + multi-run merge)
# ---------------------------------------------------------------------------
def write_upload_subset(
    parse: ParseResult,
    sku_generated: list[tuple[str, dict[str, Any]]],  # (sku, generated dict)
    path: Path,
) -> int:
    """Write an upload flat file for an arbitrary set of SKUs whose generated
    content comes from a run's results.json. Used for approving flagged SKUs
    after human review and for merging several runs into one upload file.
    Returns the number of rows written."""
    from .models import GeneratedContent

    by_sku = parse.by_sku()
    outcomes: list[SkuOutcome] = []
    for sku, gen in sku_generated:
        rec = by_sku.get(sku)
        if rec is None or not gen:
            continue
        outcomes.append(SkuOutcome(
            record=rec,
            generated=GeneratedContent(
                sku=sku,
                title=gen.get("title", ""),
                item_highlights=list(gen.get("item_highlights", [])),
                bullets=list(gen.get("bullets", [])),
                description=gen.get("description", ""),
                search_terms=gen.get("search_terms", ""),
            ),
            status="ok",
        ))
    _write_upload_xlsx(parse, outcomes, path)
    return len(outcomes)


# ---------------------------------------------------------------------------
# Bulk exports: results.csv + results.xlsx (all generated content, one row
# per SKU, ready for review in Excel/Sheets or downstream tooling)
# ---------------------------------------------------------------------------
_BULK_HEADER = [
    "sku", "product_type", "status", "review_reasons", "corrections",
    "title_before", "title_after", "title_chars",
    "highlight_1", "highlight_2", "highlight_3",
    "bullet_1_before", "bullet_1_after",
    "bullet_2_before", "bullet_2_after",
    "bullet_3_before", "bullet_3_after", "bullet_chars",
    "description_before", "description_after", "description_chars",
    "search_terms_before", "search_terms_after", "search_terms_bytes",
]


def _bulk_row(record, gen_dict: Optional[dict], status: str,
              review_reasons: str, corrections: str) -> list[Any]:
    """One export row pairing the source (before) with the generated (after).

    ``gen_dict`` uses the results.json shape so run-time outcomes and
    rebuilt-from-results exports share this single builder.
    """
    g = gen_dict or {}
    hl = list(g.get("item_highlights", [])) + ["", "", ""]
    bl = list(g.get("bullets", [])) + ["", "", ""]
    before_bl = list(record.bullets) + ["", "", ""]
    title = g.get("title", "")
    desc = g.get("description", "")
    terms = g.get("search_terms", "")
    return [
        record.sku, record.product_type, status, review_reasons, corrections,
        record.title or "", title, _chars(title),
        hl[0], hl[1], hl[2],
        before_bl[0], bl[0], before_bl[1], bl[1], before_bl[2], bl[2],
        "|".join(str(_chars(b)) for b in bl[:3]),
        record.description or "", desc, _chars(desc),
        " ".join(record.search_terms), terms, _bytes_utf8(terms),
    ]


def _bulk_rows(outcomes: list[SkuOutcome]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for o in outcomes:
        gen = o.generated
        gen_dict = None
        if gen is not None:
            gen_dict = {
                "title": gen.title, "item_highlights": list(gen.item_highlights),
                "bullets": list(gen.bullets), "description": gen.description,
                "search_terms": gen.search_terms,
            }
        reasons = "; ".join(
            f"{i.rule} ({i.field_name}): {i.message}" for i in o.issues
            if (i.severity.value if isinstance(i.severity, Severity)
                else str(i.severity)) == "error"
        ) or o.skip_reason
        clog = "; ".join(f"[{e.category}] {e.reason}" for e in o.compliance_log)
        rows.append(_bulk_row(o.record, gen_dict, o.status, reasons, clog))
    return rows


def build_bulk_rows_from_results(parse: ParseResult, results: dict) -> list[list[Any]]:
    """Rebuild export rows from a run's results.json (originals come from
    re-parsing the source file). Used by the web UI dashboard and by
    rebuilding results.csv/xlsx for runs made before the before/after export."""
    by_sku = parse.by_sku()
    rows: list[list[Any]] = []
    for s in results["skus"]:
        rec = by_sku.get(s["sku"])
        if rec is None or s["status"] not in ("ok", "needs_review", "failed"):
            continue
        reasons = "; ".join(
            f"{i['rule']} ({i['field_name']}): {i['message']}"
            for i in s.get("issues", []) if i.get("severity") == "error"
        ) or s.get("skip_reason", "")
        clog = "; ".join(f"[{e['category']}] {e['reason']}"
                         for e in s.get("compliance_log", []))
        rows.append(_bulk_row(rec, s.get("generated"), s["status"], reasons, clog))
    return rows


def _write_results_csv_rows(rows: list[list[Any]], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:  # BOM: Excel-safe
        writer = csv.writer(f)
        writer.writerow(_BULK_HEADER)
        writer.writerows(rows)


def _write_results_xlsx_rows(rows: list[list[Any]], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    header_font = openpyxl.styles.Font(bold=True)
    ws.append(_BULK_HEADER)
    for cell in ws[1]:
        cell.font = header_font
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"A": 22, "B": 16, "C": 12, "D": 45, "E": 45, "F": 45, "G": 45,
              "I": 35, "J": 35, "K": 35, "L": 40, "M": 40, "N": 40, "O": 40,
              "P": 40, "Q": 40, "S": 55, "T": 55, "V": 45, "W": 45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    wb.save(path)


# ---------------------------------------------------------------------------
# report.html
# ---------------------------------------------------------------------------
# "Aurora glass" theme: dark violet gradient canvas, frosted translucent
# cards (backdrop blur), glowing status pills. Print falls back gracefully.
_CSS = """
:root { color-scheme: dark; }
* { box-sizing: border-box; }
body { font-family: -apple-system, 'Segoe UI', Helvetica, Arial, sans-serif;
       margin: 0; padding: 32px 24px; color: rgba(255,255,255,.88);
       min-height: 100vh;
       background: linear-gradient(135deg,#0f0c29 0%,#302b63 55%,#24243e 100%)
                   fixed; }
h1 { font-size: 22px; font-weight: 600; margin: 0 0 4px; color: #fff;
     letter-spacing: .2px; }
.meta { color: rgba(255,255,255,.55); font-size: 13px; margin-bottom: 12px; }
.counts span { display: inline-block; margin: 0 8px 8px 0; padding: 4px 14px;
               border-radius: 999px; font-size: 12px; font-weight: 600;
               background: rgba(255,255,255,.08);
               border: 1px solid rgba(255,255,255,.16);
               backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px); }
.counts span.s-ok { background: rgba(52,211,153,.14);
                    border-color: rgba(52,211,153,.4); color: #6ee7b7; }
.counts span.s-needs_review { background: rgba(251,191,36,.14);
                    border-color: rgba(251,191,36,.4); color: #fcd34d; }
.counts span.s-failed { background: rgba(248,113,113,.14);
                    border-color: rgba(248,113,113,.4); color: #fca5a5; }
.card { background: rgba(255,255,255,.07); border: 1px solid rgba(255,255,255,.14);
        border-radius: 16px; padding: 18px 20px; margin: 18px 0;
        backdrop-filter: blur(14px); -webkit-backdrop-filter: blur(14px);
        box-shadow: 0 8px 32px rgba(3,0,26,.35); page-break-inside: avoid; }
.card h2 { font-size: 16px; font-weight: 600; margin: 0 0 2px; color: #fff; }
.pt { color: rgba(255,255,255,.5); font-size: 12px; margin-bottom: 10px;
      letter-spacing: .4px; }
.badge { display: inline-block; padding: 2px 10px; border-radius: 999px;
         font-size: 11px; font-weight: 600; vertical-align: middle; }
.badge.ok { background: linear-gradient(90deg,#34d399,#22d3ee); color: #052e2b; }
.badge.needs_review { background: linear-gradient(90deg,#fbbf24,#fb923c);
                      color: #451a03; }
.badge.other { background: rgba(255,255,255,.14); color: rgba(255,255,255,.75); }
.badge.error { background: rgba(248,113,113,.18); color: #fca5a5;
               border: 1px solid rgba(248,113,113,.45); }
.badge.warning { background: rgba(251,191,36,.16); color: #fcd34d;
                 border: 1px solid rgba(251,191,36,.4); }
.badge.info { background: rgba(255,255,255,.1); color: rgba(255,255,255,.65);
              border: 1px solid rgba(255,255,255,.18); }
table.ba { border-collapse: separate; border-spacing: 0; width: 100%;
           margin: 8px 0; border-radius: 10px; overflow: hidden;
           border: 1px solid rgba(255,255,255,.12); }
table.ba th, table.ba td { padding: 8px 10px; font-size: 13px; text-align: left;
           vertical-align: top; word-break: break-word;
           border-bottom: 1px solid rgba(255,255,255,.08); }
table.ba tr:last-child td { border-bottom: none; }
table.ba th { background: rgba(255,255,255,.08); color: rgba(255,255,255,.75);
              font-weight: 600; font-size: 12px; }
table.ba td { background: rgba(255,255,255,.03); color: rgba(255,255,255,.82); }
table.ba td.field { white-space: nowrap; font-weight: 600; width: 110px;
                    color: #a5b4fc; }
.count { color: rgba(255,255,255,.45); font-size: 11px; white-space: nowrap; }
.count.over { color: #fca5a5; font-weight: 700; }
.hl-note { color: #fcd34d; font-size: 12px; margin: 2px 0 4px; }
ul.issues, ul.hl, ul.clog { margin: 4px 0 0; padding-left: 18px; }
ul.issues li, ul.clog li { font-size: 13px; margin: 4px 0;
                           color: rgba(255,255,255,.78); }
ul.hl li { font-size: 13px; margin: 3px 0; color: rgba(255,255,255,.85); }
.rule { font-family: ui-monospace, Menlo, monospace; font-size: 12px;
        color: #93c5fd; }
h3 { font-size: 13px; font-weight: 600; margin: 14px 0 4px;
     color: rgba(255,255,255,.65); letter-spacing: .5px;
     text-transform: uppercase; }
.empty { color: rgba(255,255,255,.3); font-style: italic; }
@media print {
  body { background: #fff; color: #111; }
  .card { background: #fff; border-color: #ddd; box-shadow: none; }
}
"""


def _esc(value: Optional[str]) -> str:
    if not value:
        return ""
    return html.escape(str(value))


def _cell(value: Optional[str]) -> str:
    return _esc(value) if value else '<span class="empty">—</span>'


def _count_span(n: int, limit: int, unit: str) -> str:
    cls = "count over" if n > limit else "count"
    return f'<span class="{cls}">{n}/{limit} {unit}</span>'


def _status_badge(status: str) -> str:
    cls = status if status in ("ok", "needs_review") else "other"
    return f'<span class="badge {cls}">{_esc(status)}</span>'


def _ba_row(field: str, before: Optional[str], after: Optional[str],
            count_html: str) -> str:
    return (
        f'<tr><td class="field">{_esc(field)}</td>'
        f"<td>{_cell(before)}</td>"
        f"<td>{_cell(after)} {count_html}</td></tr>"
    )


def _card_html(outcome: SkuOutcome) -> str:
    rec = outcome.record
    gen = outcome.generated
    parts: list[str] = ['<div class="card">']
    parts.append(
        f"<h2>{_esc(rec.sku)} {_status_badge(outcome.status)}</h2>"
        f'<div class="pt">{_esc(rec.product_type)}'
        + (f" — {_esc(outcome.skip_reason)}" if outcome.skip_reason else "")
        + "</div>"
    )

    # Before/After table
    parts.append(
        '<table class="ba"><tr><th>Field</th><th>Before (source)</th>'
        "<th>After (generated)</th></tr>"
    )
    title_after = gen.title if gen else ""
    parts.append(
        _ba_row("Title", rec.title, title_after,
                _count_span(_chars(title_after), TITLE_MAX_CHARS, "chars"))
    )
    before_bullets = rec.bullets
    after_bullets = gen.bullets if gen else []
    for i in range(3):
        before = before_bullets[i] if i < len(before_bullets) else ""
        after = after_bullets[i] if i < len(after_bullets) else ""
        parts.append(
            _ba_row(f"Bullet {i + 1}", before, after,
                    _count_span(_chars(after), BULLET_MAX_CHARS, "chars"))
        )
    desc_after = gen.description if gen else ""
    parts.append(
        _ba_row("Description", rec.description, desc_after,
                _count_span(_chars(desc_after), DESCRIPTION_MAX_CHARS, "chars"))
    )
    terms_after = gen.search_terms if gen else ""
    parts.append(
        _ba_row("Search terms", " ".join(rec.search_terms), terms_after,
                _count_span(_bytes_utf8(terms_after), SEARCH_TERMS_MAX_BYTES,
                            "bytes"))
    )
    parts.append("</table>")

    # Item highlights (after only — no template column exists for these)
    if gen and gen.item_highlights:
        parts.append("<h3>Item Highlights</h3>")
        parts.append(
            '<div class="hl-note">not uploaded — paste into Seller Central '
            "manually</div>"
        )
        parts.append('<ul class="hl">')
        for hl in gen.item_highlights:
            parts.append(
                f"<li>{_esc(hl)} "
                f"{_count_span(_chars(hl), HIGHLIGHT_MAX_CHARS, 'chars')}</li>"
            )
        parts.append("</ul>")

    # Issues
    if outcome.issues:
        parts.append("<h3>Issues</h3>")
        parts.append('<ul class="issues">')
        for issue in outcome.issues:
            sev = issue.severity.value if isinstance(issue.severity, Severity) \
                else str(issue.severity)
            parts.append(
                f'<li><span class="badge {_esc(sev)}">{_esc(sev)}</span> '
                f'<span class="rule">{_esc(issue.rule)}</span> '
                f"({_esc(issue.field_name)}): {_esc(issue.message)}</li>"
            )
        parts.append("</ul>")

    # Compliance log
    if outcome.compliance_log:
        parts.append("<h3>Compliance log</h3>")
        parts.append('<ul class="clog">')
        for entry in outcome.compliance_log:
            parts.append(
                f'<li><span class="rule">{_esc(entry.field_name)}</span> '
                f"[{_esc(entry.category)}]: {_esc(entry.reason)}</li>"
            )
        parts.append("</ul>")

    parts.append("</div>")
    return "".join(parts)


def _write_report_html(
    parse: ParseResult, outcomes: list[SkuOutcome], run_id: str, path: Path
) -> None:
    counts = Counter(o.status for o in outcomes)
    counts_html = "".join(
        f'<span class="s-{_esc(status)}">{_esc(status)}: {n}</span>'
        for status, n in sorted(counts.items())
    )
    ordered = sorted(
        outcomes, key=lambda o: (_STATUS_ORDER.get(o.status, 2), o.record.sku)
    )
    cards = "".join(_card_html(o) for o in ordered)
    doc = (
        "<!DOCTYPE html>\n"
        '<html lang="en"><head><meta charset="utf-8">'
        f"<title>Catalog run {_esc(run_id)}</title>"
        f"<style>{_CSS}</style></head><body>"
        f"<h1>Catalog optimization report</h1>"
        f'<div class="meta">Run {_esc(run_id)} — source '
        f"{_esc(Path(parse.source_file).name)} — generated "
        f"{_esc(datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'))}</div>"
        f'<div class="counts">{counts_html}</div>'
        f"{cards}"
        "</body></html>"
    )
    path.write_text(doc, encoding="utf-8")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def write_outputs(
    parse: ParseResult,
    outcomes: list[SkuOutcome],
    run_id: str,
    out_dir: Path,
) -> dict[str, str]:
    """Write all run artifacts into out_dir; return artifact name -> path.

    Artifacts: upload.xlsx (Amazon partial-update flat file, ok SKUs only),
    results.json (full per-SKU results incl. item highlights and counts),
    report.html (self-contained human review report), input_issues.json
    (input-data issues, field_name starting 'input:').
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    ok_outcomes = [o for o in outcomes if o.status == "ok"]

    upload_path = out_dir / "upload.xlsx"
    results_path = out_dir / "results.json"
    report_path = out_dir / "report.html"
    input_issues_path = out_dir / "input_issues.json"
    csv_path = out_dir / "results.csv"
    xlsx_path = out_dir / "results.xlsx"

    reviewable = [o for o in outcomes if o.status in ("ok", "needs_review", "failed")]

    bulk = _bulk_rows(reviewable)
    _write_upload_xlsx(parse, ok_outcomes, upload_path)
    _write_results_json(parse, outcomes, run_id, results_path)
    _write_report_html(parse, outcomes, run_id, report_path)
    _write_input_issues_json(outcomes, input_issues_path)
    _write_results_csv_rows(bulk, csv_path)
    _write_results_xlsx_rows(bulk, xlsx_path)
    # precomputed dashboard rows — web UI loads this instead of re-parsing
    _dump_json({"header": _BULK_HEADER, "rows": bulk}, out_dir / "rows.json")

    return {
        "upload.xlsx": str(upload_path),
        "results.json": str(results_path),
        "results.csv": str(csv_path),
        "results.xlsx": str(xlsx_path),
        "report.html": str(report_path),
        "input_issues.json": str(input_issues_path),
    }
