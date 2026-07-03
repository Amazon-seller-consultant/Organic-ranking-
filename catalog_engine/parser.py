"""Category Listing Report ingestion.

Reads Amazon's .xlsm/.xlsx flat file dynamically:
- header row positions come from the settings string in Template!A1
  (labelRow / attributeRow / dataRow) — never hardcoded;
- columns are mapped by parsing the machine attribute names on attributeRow,
  because column positions shift between exports (verified across the three
  sample files: 812 / 925 / 658 columns, same attributes, different indices);
- malformed uploads fail with a specific seller-facing FlatFileError.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Optional
from urllib.parse import parse_qs, unquote

import openpyxl

from .exceptions import FlatFileError
from .models import (
    ColumnInfo,
    Parentage,
    ParseResult,
    ProductRecord,
    RecordAction,
    TemplateSettings,
    VariationFamily,
)

TEMPLATE_SHEET = "Template"
REQUIRED_BASES = {"contribution_sku", "product_type", "::record_action"}

# amzn marketplace obfuscated ids -> short codes (extend as sellers onboard)
MARKETPLACE_CODES = {
    "ATVPDKIKX0DER": "US",
    "A1F83G8C2ARO7P": "UK",
    "A13V1IB3VIYZZH": "FR",
    "A1PA6795UKMFR9": "DE",
    "APJ6JRA9NG5V4": "IT",
    "A1RKKUPIHCS9HS": "ES",
    "A2EUQ1WTGCTBG2": "CA",
}

_QUAL_RE = re.compile(r"\[[^\]]*\]")
_INSTANCE_RE = re.compile(r"#(\d+)")


def parse_column_header(index: int, raw_name: str, label: str = "") -> ColumnInfo:
    """Parse a machine attribute name like
    bullet_point[marketplace_id=X][language_tag=Y]#2.value  or
    light_source[...]#1.special_features[...]#2.value  or
    ::record_action
    into (base, instance, leaf).
    """
    stripped = _QUAL_RE.sub("", raw_name.strip())
    instances = _INSTANCE_RE.findall(stripped)
    instance = int(instances[-1]) if instances else 1
    path = _INSTANCE_RE.sub("", stripped)  # e.g. light_source.special_features.value
    parts = path.split(".")
    if len(parts) > 1:
        leaf = parts[-1]
        base = ".".join(parts[:-1])
    else:
        leaf, base = "", parts[0]
    return ColumnInfo(
        index=index, raw_name=raw_name, base=base, instance=instance,
        leaf=leaf, label=label or "",
    )


def attribute_key(col: ColumnInfo) -> str:
    """Key under which this column's values land in ProductRecord.attributes."""
    if col.leaf in ("value", ""):
        return col.base
    return f"{col.base}.{col.leaf}"


def parse_settings(a1_value: Optional[str], source: str) -> TemplateSettings:
    if not a1_value or not str(a1_value).startswith("settings="):
        raise FlatFileError(
            f"'{source}': cell A1 of the {TEMPLATE_SHEET} sheet does not contain the "
            "expected Amazon settings string (should start with 'settings=').",
            "Re-download the Category Listing Report from Seller Central and upload "
            "it unmodified.",
        )
    qs = parse_qs(str(a1_value).split("settings=", 1)[1])

    def need_int(key: str) -> int:
        if key not in qs:
            raise FlatFileError(
                f"'{source}': the settings string in {TEMPLATE_SHEET}!A1 is missing "
                f"'{key}', so header rows cannot be located.",
                "This file may be from an unsupported template version.",
            )
        return int(qs[key][0])

    raw_mp = unquote(qs.get("primaryMarketplaceId", [""])[0])
    mp_token = raw_mp.rsplit(".", 1)[-1] if raw_mp else ""
    return TemplateSettings(
        label_row=need_int("labelRow"),
        attribute_row=need_int("attributeRow"),
        data_row=need_int("dataRow"),
        marketplace_id=MARKETPLACE_CODES.get(mp_token, mp_token or "UNKNOWN"),
        language_tag=qs.get("contentLanguageTag", ["en_US"])[0],
        template_identifier=qs.get("templateIdentifier", [""])[0],
        feed_type=qs.get("feedType", [""])[0],
        raw={k: v[0] for k, v in qs.items()},
    )


def _load_workbook(path: Path):
    if not path.exists():
        raise FlatFileError(f"file not found: {path}")
    if path.suffix.lower() not in (".xlsm", ".xlsx"):
        raise FlatFileError(
            f"'{path.name}' is a {path.suffix or 'file with no extension'} file, "
            "not an Excel workbook.",
            "Upload the Category Listing Report exactly as downloaded from Seller "
            "Central (.xlsm).",
        )
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (zipfile.BadZipFile, KeyError, OSError) as exc:
        raise FlatFileError(
            f"'{path.name}' could not be opened as an Excel workbook "
            f"({type(exc).__name__}). The file appears to be corrupted or truncated.",
            "Re-download it from Seller Central and try again.",
        ) from exc


def parse_flat_file(path: str | Path, seller_id: str) -> ParseResult:
    path = Path(path)
    wb = _load_workbook(path)
    try:
        if TEMPLATE_SHEET not in wb.sheetnames:
            raise FlatFileError(
                f"'{path.name}' has no '{TEMPLATE_SHEET}' sheet "
                f"(found: {', '.join(wb.sheetnames) or 'none'}).",
                "This does not look like an Amazon Category Listing Report.",
            )
        ws = wb[TEMPLATE_SHEET]
        settings = parse_settings(ws.cell(1, 1).value, path.name)

        rows_needed = max(settings.label_row, settings.attribute_row)
        header_rows: dict[int, tuple] = {}
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=rows_needed, values_only=True), 1
        ):
            header_rows[i] = row
        attr_row = header_rows.get(settings.attribute_row) or ()
        label_row = header_rows.get(settings.label_row) or ()
        if not any(attr_row):
            raise FlatFileError(
                f"'{path.name}': row {settings.attribute_row} (attributeRow per the "
                "settings string) is empty — the template structure does not match "
                "its own settings.",
            )

        columns: list[ColumnInfo] = []
        for idx, raw_name in enumerate(attr_row):
            if raw_name is None or not str(raw_name).strip():
                continue
            label = ""
            if idx < len(label_row) and label_row[idx] is not None:
                label = str(label_row[idx])
            columns.append(parse_column_header(idx, str(raw_name), label))

        bases = {c.base for c in columns}
        missing = REQUIRED_BASES - bases
        if missing:
            raise FlatFileError(
                f"'{path.name}': required column(s) not found on the attribute row: "
                f"{', '.join(sorted(missing))}.",
                "The template structure may have changed; the parser needs updating "
                "before this file can be processed safely.",
            )

        # Column lookup helpers
        def cols_for(base: str, leaf: str = "") -> list[ColumnInfo]:
            found = [
                c for c in columns
                if c.base == base and (leaf == "" or c.leaf == leaf)
            ]
            return sorted(found, key=lambda c: c.instance)

        sku_col = cols_for("contribution_sku")[0]
        pt_col = cols_for("product_type")[0]
        action_col = cols_for("::record_action")[0]
        status_cols = cols_for("::listing_status")
        parentage_cols = cols_for("parentage_level")
        parent_sku_cols = cols_for("child_parent_sku_relationship", "parent_sku")
        theme_cols = cols_for("variation_theme")

        warnings: list[str] = []
        records: list[ProductRecord] = []
        seen_skus: set[str] = set()

        for row_no, row in enumerate(
            ws.iter_rows(min_row=settings.data_row, values_only=True),
            settings.data_row,
        ):
            if not any(v is not None and str(v).strip() for v in row):
                continue

            def cell(col: Optional[ColumnInfo]) -> str:
                if col is None or col.index >= len(row) or row[col.index] is None:
                    return ""
                return str(row[col.index]).strip()

            sku = cell(sku_col)
            if not sku:
                warnings.append(
                    f"row {row_no}: has data but no SKU (contribution_sku) — skipped"
                )
                continue
            if sku in seen_skus:
                warnings.append(f"row {row_no}: duplicate SKU '{sku}' — keeping first")
                continue
            seen_skus.add(sku)

            action = RecordAction.from_cell(cell(action_col) or None)
            if action is RecordAction.UNKNOWN:
                warnings.append(
                    f"row {row_no} (SKU {sku}): unrecognized record action "
                    f"'{cell(action_col)}' — flagged, not guessed"
                )

            parentage_raw = cell(parentage_cols[0]) if parentage_cols else ""
            if parentage_raw.lower() == "parent":
                parentage = Parentage.PARENT
            elif parentage_raw.lower() == "child":
                parentage = Parentage.CHILD
            else:
                parentage = Parentage.STANDALONE

            attributes: dict[str, list[tuple[int, str]]] = {}
            raw_cells: dict[str, str] = {}
            for col in columns:
                val = cell(col)
                if not val:
                    continue
                raw_cells[col.raw_name] = val
                attributes.setdefault(attribute_key(col), []).append(
                    (col.instance, val)
                )

            records.append(
                ProductRecord(
                    seller_id=seller_id,
                    sku=sku,
                    row_number=row_no,
                    product_type=cell(pt_col),
                    record_action=action,
                    listing_status=cell(status_cols[0]) if status_cols else "",
                    parentage=parentage,
                    parent_sku=(cell(parent_sku_cols[0]) or None)
                    if parent_sku_cols else None,
                    variation_theme=(cell(theme_cols[0]) or None)
                    if theme_cols else None,
                    attributes={
                        k: [v for _, v in sorted(vals)] for k, vals in attributes.items()
                    },
                    raw_cells=raw_cells,
                )
            )

        if not records:
            raise FlatFileError(
                f"'{path.name}': no product rows found at or after row "
                f"{settings.data_row} (dataRow per the settings string).",
                "The report appears to be empty.",
            )

        families = group_families(records, warnings)
        return ParseResult(
            seller_id=seller_id,
            source_file=str(path),
            settings=settings,
            columns=columns,
            records=records,
            families=families,
            warnings=warnings,
        )
    finally:
        wb.close()


def group_families(
    records: list[ProductRecord], warnings: list[str]
) -> list[VariationFamily]:
    by_sku = {r.sku: r for r in records}
    parents = {r.sku: r for r in records if r.parentage is Parentage.PARENT}
    fams: dict[str, VariationFamily] = {
        sku: VariationFamily(parent_sku=sku, parent=rec) for sku, rec in parents.items()
    }
    for rec in records:
        if rec.parentage is not Parentage.CHILD:
            continue
        psku = rec.parent_sku or ""
        if not psku:
            warnings.append(
                f"SKU {rec.sku}: marked Child but has no parent SKU — orphan child"
            )
            fams.setdefault(
                "", VariationFamily(parent_sku="", parent=None)
            ).children.append(rec)
            continue
        if psku not in fams:
            if psku in by_sku:
                # parent row exists but is not flagged Parent — group anyway, warn
                warnings.append(
                    f"SKU {rec.sku}: parent '{psku}' exists but its parentage_level "
                    "is not 'Parent'"
                )
                fams[psku] = VariationFamily(parent_sku=psku, parent=by_sku[psku])
            else:
                warnings.append(
                    f"SKU {rec.sku}: parent SKU '{psku}' not present in this report — "
                    "orphan child"
                )
                fams[psku] = VariationFamily(parent_sku=psku, parent=None)
        fams[psku].children.append(rec)
    return list(fams.values())
